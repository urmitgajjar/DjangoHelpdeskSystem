from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.views import PasswordResetView
from django.contrib import messages
from django.db.models import Q, Avg, OuterRef, Subquery, Count
from django.http import HttpResponse, FileResponse, JsonResponse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.paginator import Paginator
from django.urls import reverse
from datetime import datetime, time
from myapp.ml_models.department_predictor import predict_department
from myapp.models import Department
import json
import logging
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io

from .models import (
    UserProfile, TicketDetail, MyCart, ActivityLog,
    UserComment, Category, Notification, Department, DepartmentMember,
    TicketHistory, TicketRating, AIMLLog,
    CannedResponse,
    get_visible_active_memberships,
)
from .decorators import (
    department_member_required,
    admin_required,
    LoginRoleAuthorization,
    ROLE_PERMISSION_MATRIX,
    user_has_department_permission,
)
from .analytics import (
    get_date_range, get_ticket_statistics, get_tickets_over_time,
    get_department_statistics, get_department_comparison,
    get_top_ticket_creators, get_top_ticket_resolvers,
    get_priority_distribution, get_category_distribution,
    prepare_export_data,
)
from .notifications import (
    create_notification,
    notify_ticket_created, notify_ticket_updated,
    notify_ticket_due_date_extended,
    notify_ticket_closed, notify_ticket_resolved, notify_ticket_reopened,
    notify_ticket_commented,
    notify_ticket_rated,
)
from .ai.ai_priority import predict_ticket_priority_with_meta
from .forms import (
    LoginForm, RegisterForm, UserProfileForm, TicketDetailForm, TicketCreateForm,
    UserCommentForm, TicketUpdateForm, TicketFilterForm, CategoryForm,
    AccountSettingsForm, UsernameEmailPasswordResetForm, AdminTicketRoutingForm,
    DepartmentMemberForm, DepartmentAdminForm, TicketRatingForm,
)


logger = logging.getLogger(__name__)


def log_activity(user, action, title, description='', ticket=None, old_value='', new_value=''):
    ActivityLog.objects.create(
        user=user, ticket=ticket, action=action,
        title=title, description=description,
        old_value=old_value, new_value=new_value,
    )


def _log_priority_prediction(ticket, prediction: dict):
    priority = (prediction or {}).get("priority")
    if not priority:
        return

    AIMLLog.objects.create(
        ticket=ticket,
        log_type='PRIORITY',
        input_data=json.dumps({
            "title": ticket.TICKET_TITLE,
            "description": ticket.TICKET_DESCRIPTION,
        }),
        output_data=json.dumps({
            "priority": priority,
            "reason": (prediction or {}).get("reason", ""),
            "model": (prediction or {}).get("model", ""),
            "error": (prediction or {}).get("error", ""),
        }),
        confidence=None,
        was_correct=None,
    )


def _record_priority_feedback(ticket, selected_priority, user):
    suggested = (ticket.ai_suggested_priority or '').strip().upper()
    selected = (selected_priority or '').strip().upper()
    if not suggested or not selected:
        return

    latest_log = AIMLLog.objects.filter(ticket=ticket, log_type='PRIORITY').order_by('-created_at').first()
    if not latest_log:
        return

    try:
        output_payload = json.loads(latest_log.output_data or "{}")
    except json.JSONDecodeError:
        output_payload = {}

    output_payload["user_selected_priority"] = selected
    output_payload["suggested_priority"] = suggested
    output_payload["feedback_by"] = getattr(user, "username", "")
    latest_log.output_data = json.dumps(output_payload)
    latest_log.was_correct = (selected == suggested)
    latest_log.save(update_fields=['output_data', 'was_correct'])


def _is_admin_user(user):
    if not user or not user.is_authenticated:
        return False
    return user.is_superuser

def _ensure_userprofile_and_permissions(user):
    if not user or not user.is_authenticated:
        return None
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


def _is_department_member(user, ticket):
    if not ticket.assigned_department:
        return False
    return DepartmentMember.objects.filter(
        user=user,
        department=ticket.assigned_department,
        is_active=True,
        department__is_active=True,
    ).exists()


def _can_view_ticket(user, ticket):
    if _is_admin_user(user):
        return True
    if ticket.__class__.objects.filter(id=ticket.id).filter(_inactive_department_ticket_q()).exists():
        return False
    if ticket.TICKET_CREATED_id == user.id:
        return True
    if ticket.assigned_to_id == user.id:
        return True
                                                                            
    return MyCart.objects.filter(user=user, ticket=ticket).exists()

def _can_work_on_ticket(user, ticket):
    if _is_admin_user(user):
        return False
    if ticket.assigned_department_id:
        dept = ticket.assigned_department
        if not getattr(dept, 'is_active', True):
            return False
    if ticket.TICKET_CREATED_id == user.id:
        return True
    if ticket.assigned_to_id == user.id:
        return True
    return MyCart.objects.filter(user=user, ticket=ticket).exists()


def _is_close_locked_by_rejection(user, ticket):
    latest_rejection = (
        TicketHistory.objects
        .filter(ticket=ticket, action_type='REJECTED', changed_by=user)
        .order_by('-changed_at')
        .first()
    )
    if not latest_rejection:
        return False

    reassigned_after_rejection = TicketHistory.objects.filter(
        ticket=ticket,
        action_type='ASSIGNED',
        new_value=user.username,
        changed_at__gt=latest_rejection.changed_at,
    ).exists()
    return not reassigned_after_rejection


def _can_user_close_ticket(user, ticket):
    if _is_admin_user(user):
        return False
    if ticket.TICKET_STATUS in ['Closed', 'Resolved']:
        return False
    if not _can_work_on_ticket(user, ticket):
        return False
    return not _is_close_locked_by_rejection(user, ticket)


def _is_single_member_department_assignment(ticket):
    if not ticket.assigned_department_id or not getattr(ticket.assigned_department, 'is_active', False):
        return False
    active_member_count = DepartmentMember.objects.filter(
        department=ticket.assigned_department,
        is_active=True,
        user__is_active=True,
    ).count()
    return active_member_count <= 1


def _auto_assign_single_member_department_ticket(ticket, changed_by=None):
    if not ticket.assigned_department_id or ticket.assigned_to_id:
        return None
    if not getattr(ticket.assigned_department, 'is_active', False):
        return None
    memberships = (
        DepartmentMember.objects
        .filter(
            department=ticket.assigned_department,
            is_active=True,
            user__is_active=True,
        )
        .select_related('user')
    )
    membership_list = list(memberships[:2])
    if len(membership_list) != 1:
        return None

    assignee = membership_list[0].user
    ticket.assigned_to = assignee
    ticket.TICKET_HOLDER = assignee.username
    if changed_by and not ticket.assigned_by_id:
        ticket.assigned_by = changed_by
    if not ticket.assigned_at:
        ticket.assigned_at = timezone.now()
    ticket.save(update_fields=['assigned_to', 'TICKET_HOLDER', 'assigned_by', 'assigned_at'])

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=changed_by,
        action_type='ASSIGNED',
        old_value='Unassigned',
        new_value=assignee.username,
        description=(
            f'Auto-assigned to {assignee.username} because they are the only active member '
            f'in {ticket.assigned_department.name}.'
        ),
    )
    MyCart.objects.get_or_create(user=assignee, ticket=ticket)
    return assignee


def _is_creator_only_member_of_department(user, department):
    if not user or not department:
        return False
    active_memberships = DepartmentMember.objects.filter(
        department=department,
        is_active=True,
        user__is_active=True,
    )
    return active_memberships.count() == 1 and active_memberships.filter(user=user).exists()


def _restore_department_ticket_assignments(department, changed_by=None):
    if not department or not department.is_active:
        return 0

    restored_count = 0
    tickets = TicketDetail.objects.filter(
        assigned_department=department,
        assigned_to__isnull=True,
    ).select_related('assigned_department')

    for ticket in tickets:
        history_entry = (
            TicketHistory.objects.filter(
                ticket=ticket,
                field_name='department_inactivation_assignee',
            )
            .exclude(old_value='')
            .order_by('-changed_at')
            .first()
        )
        if not history_entry:
            _auto_assign_single_member_department_ticket(ticket, changed_by=changed_by)
            continue

        try:
            previous_assignee_id = int(history_entry.old_value)
        except (TypeError, ValueError):
            previous_assignee_id = None

        if not previous_assignee_id:
            _auto_assign_single_member_department_ticket(ticket, changed_by=changed_by)
            continue

        membership = (
            DepartmentMember.objects.filter(
                user_id=previous_assignee_id,
                department=department,
                is_active=True,
                department__is_active=True,
                user__is_active=True,
            )
            .select_related('user')
            .first()
        )
        if not membership:
            _auto_assign_single_member_department_ticket(ticket, changed_by=changed_by)
            continue

        ticket.assigned_to = membership.user
        ticket.TICKET_HOLDER = membership.user.username
        ticket.assignment_type = 'MANUAL'
        ticket.assigned_by = changed_by
        ticket.assigned_at = timezone.now()
        ticket.save(update_fields=['assigned_to', 'TICKET_HOLDER', 'assignment_type', 'assigned_by', 'assigned_at'])
        MyCart.objects.get_or_create(user=membership.user, ticket=ticket)

        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=changed_by,
            action_type='ASSIGNED',
            field_name='assigned_to',
            old_value='Unassigned',
            new_value=membership.user.username,
            description=(
                f'Assignment restored to {membership.user.username} after '
                f'{department.name} department was reactivated.'
            ),
        )
        restored_count += 1

    return restored_count


def _sync_mycart_for_user(user):
    if not user.is_authenticated or user.is_superuser:
        return

    department_ids = DepartmentMember.objects.filter(
        user=user, is_active=True, department__is_active=True
    ).values_list('department_id', flat=True)

    rejected_ticket_ids = set(
        TicketHistory.objects.filter(
            changed_by=user,
            action_type='REJECTED'
        ).values_list('ticket_id', flat=True)
    )

    eligible_tickets = TicketDetail.objects.filter(
        ~Q(TICKET_STATUS__in=['Closed', 'Resolved'])
    ).filter(
        Q(assigned_to=user) |
        (
            Q(assigned_department_id__in=department_ids) &
            Q(assigned_to__isnull=True)
        )
    ).exclude(
        TICKET_CREATED=user
    ).exclude(
        Q(id__in=rejected_ticket_ids) & ~Q(assigned_to=user)
    ).distinct()

    eligible_ticket_ids = set(eligible_tickets.values_list('id', flat=True))
    existing_ticket_ids = set(
        MyCart.objects.filter(user=user).values_list('ticket_id', flat=True)
    )

    for ticket_id in (eligible_ticket_ids - existing_ticket_ids):
        MyCart.objects.create(user=user, ticket_id=ticket_id)

    if existing_ticket_ids:
        MyCart.objects.filter(user=user).exclude(
            ticket_id__in=eligible_ticket_ids
        ).delete()

    if rejected_ticket_ids:
        MyCart.objects.filter(
            user=user,
            ticket_id__in=rejected_ticket_ids
        ).exclude(ticket__assigned_to=user).delete()


def _is_non_rejectable_assignment(user, ticket):
    if ticket.TICKET_STATUS == 'Reopen':
        return True

    if ticket.assigned_by_id and ticket.assigned_by.is_superuser:
        return True

    latest_assigned = (
        TicketHistory.objects
        .filter(ticket=ticket, action_type='ASSIGNED')
        .order_by('-changed_at')
        .first()
    )
    if (
        latest_assigned
        and latest_assigned.new_value == user.username
        and latest_assigned.description
        and 'Auto-assigned to' in latest_assigned.description
    ):
        latest_reopened = (
            TicketHistory.objects
            .filter(ticket=ticket, action_type='REOPENED')
            .order_by('-changed_at')
            .first()
        )
        if not latest_reopened or latest_reopened.changed_at <= latest_assigned.changed_at:
            return True

    return False


def _overdue_note_thread_qs(ticket):
    return (
        TicketHistory.objects
        .filter(
            ticket=ticket,
            field_name__in=['admin_overdue_note', 'admin_overdue_note_reply'],
        )
        .select_related('changed_by')
        .order_by('changed_at')
    )


def _ticketinfo_overdue_redirect(ticket):
    paginator = Paginator(_overdue_note_thread_qs(ticket), 4)
    page_number = paginator.num_pages or 1
    return redirect(f"{reverse('ticketinfo', kwargs={'pk': ticket.id})}?overdue_page={page_number}#overdue-note-thread")


def _auto_assign_on_department_rejection(ticket, rejected_by):
    if not ticket.assigned_department_id:
        return None
    if not getattr(ticket.assigned_department, 'is_active', False):
        return None
    if ticket.TICKET_STATUS in ['Closed', 'Resolved']:
        return None

    memberships = (
        DepartmentMember.objects
        .filter(
            department=ticket.assigned_department,
            is_active=True,
            user__is_active=True,
        )
        .select_related('user')
        .order_by('-id')
    )
    if not memberships.exists():
        return None

    creator_id = ticket.TICKET_CREATED_id
    assignable_memberships = memberships.exclude(user_id=creator_id)
    if not assignable_memberships.exists():
        return None

    preferred_memberships = assignable_memberships.exclude(user_id=rejected_by.id)
    active_user_ids = list(assignable_memberships.values_list('user_id', flat=True))
    rejected_user_ids = set(
        TicketHistory.objects.filter(
            ticket=ticket,
            action_type='REJECTED',
            changed_by_id__in=active_user_ids,
        ).values_list('changed_by_id', flat=True)
    )
    rejected_user_ids.add(rejected_by.id)

    remaining_memberships = preferred_memberships.exclude(user_id__in=rejected_user_ids)
    remaining_count = remaining_memberships.count()

    assignee_membership = None
    if remaining_count >= 1:
        assignee_membership = remaining_memberships.first()

    if not assignee_membership:
        return None

    assignee = assignee_membership.user
    if ticket.assigned_to_id == assignee.id:
        return assignee

    old_assignee_name = ticket.assigned_to.username if ticket.assigned_to_id else 'Unassigned'
    ticket.assigned_to = assignee
    ticket.TICKET_HOLDER = assignee.username
    ticket.save(update_fields=['assigned_to', 'TICKET_HOLDER'])
    MyCart.objects.get_or_create(user=assignee, ticket=ticket)

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=rejected_by,
        action_type='ASSIGNED',
        old_value=old_assignee_name,
        new_value=assignee.username,
        description=(
            f'Auto-assigned to {assignee.username} after department rejections.'
        ),
    )
    create_notification(
        user=assignee,
        notification_type='TICKET_ASSIGNED',
        title=f'Auto-assigned ticket #{ticket.id}',
        message=f'Ticket "{ticket.TICKET_TITLE}" was auto-assigned to you after rejections.',
        ticket=ticket,
        extra_data={
            'auto_assigned': True,
            'department': ticket.assigned_department.name if ticket.assigned_department else '',
            'triggered_by': rejected_by.username,
        },
    )
    return assignee


def _department_member_workload_rows(ticket):
    if not ticket.assigned_department_id or not getattr(ticket.assigned_department, 'is_active', False):
        return []

    memberships = list(
        DepartmentMember.objects
        .filter(
            department=ticket.assigned_department,
            is_active=True,
            department__is_active=True,
            user__is_active=True,
        )
        .exclude(user_id=ticket.TICKET_CREATED_id)
        .select_related('user')
        .order_by('user__username')
    )
    if not memberships:
        return []

    user_ids = [membership.user_id for membership in memberships]
    active_tickets = TicketDetail.objects.exclude(TICKET_STATUS__in=['Closed', 'Resolved'])

    total_counts = {
        row['assigned_to']: row['total']
        for row in active_tickets.filter(assigned_to_id__in=user_ids)
        .values('assigned_to')
        .annotate(total=Count('id'))
    }
    department_counts = {
        row['assigned_to']: row['total']
        for row in active_tickets.filter(
            assigned_to_id__in=user_ids,
            assigned_department=ticket.assigned_department,
        )
        .values('assigned_to')
        .annotate(total=Count('id'))
    }
    rejected_user_ids = set(
        TicketHistory.objects.filter(
            ticket=ticket,
            action_type='REJECTED',
            changed_by_id__in=user_ids,
        ).values_list('changed_by_id', flat=True)
    )

    rows = []
    for membership in memberships:
        rows.append({
            'user': membership.user,
            'role': membership.display_role,
            'same_department_open_count': department_counts.get(membership.user_id, 0),
            'total_open_count': total_counts.get(membership.user_id, 0),
            'has_rejected_ticket': membership.user_id in rejected_user_ids,
            'is_current_assignee': membership.user_id == ticket.assigned_to_id,
        })

    rows.sort(
        key=lambda row: (
            row['is_current_assignee'],
            row['same_department_open_count'],
            row['total_open_count'],
            row['user'].username.lower(),
        )
    )
    return rows


def _eligible_reassignment_memberships(ticket):
    if not ticket.assigned_department_id or not getattr(ticket.assigned_department, 'is_active', False):
        return DepartmentMember.objects.none()
    return (
        DepartmentMember.objects
        .filter(
            department=ticket.assigned_department,
            is_active=True,
            department__is_active=True,
            user__is_active=True,
        )
        .exclude(user_id=ticket.TICKET_CREATED_id)
        .select_related('user')
    )


def _get_primary_department_id(user):
    membership = DepartmentMember.objects.filter(
        user=user, is_active=True, department__is_active=True
    ).order_by('department__name').first()
    return membership.department_id if membership else None


def _inactive_department_member_user_ids():
    return DepartmentMember.objects.filter(
        is_active=True,
        user__is_active=True,
        department__is_active=False,
    ).values('user_id')


def _inactive_department_ticket_q(prefix=''):
    return (
        Q(**{f'{prefix}assigned_department__is_active': False}) |
        Q(**{f'{prefix}TICKET_CREATED_id__in': _inactive_department_member_user_ids()})
    )


def _exclude_inactive_department_tickets(queryset, prefix=''):
    return queryset.exclude(_inactive_department_ticket_q(prefix))


def _user_has_inactive_department_membership(user):
    if not user or not user.is_authenticated or user.is_superuser:
        return False
    return DepartmentMember.objects.filter(
        user=user,
        is_active=True,
        department__is_active=False,
    ).exists()


def _get_dashboard_redirect_url(user):
    if _is_admin_user(user):
        return reverse('base')
    return reverse('base')


class UsernameRequiredPasswordResetView(PasswordResetView):
    form_class = UsernameEmailPasswordResetForm

    def _requested_username(self):
        return (self.request.POST.get('username') or self.request.GET.get('username') or '').strip()

    def dispatch(self, request, *args, **kwargs):
        requested_username = self._requested_username()
        if not requested_username:
            return redirect('login')
        if not User.objects.filter(username__iexact=requested_username, is_active=True).exists():
            messages.error(request, "Username not found. Enter a valid registered username.")
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['username'] = self._requested_username()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['requested_username'] = self._requested_username()
        return context


def username_exists_api(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    username = (request.GET.get('username') or '').strip()
    exists = False
    if username:
        exists = User.objects.filter(username__iexact=username, is_active=True).exists()
    return JsonResponse({'exists': exists})

def _notifications_for_user(user):
    base_qs = Notification.objects.filter(user=user).select_related('ticket', 'ticket__assigned_department')
    if _is_admin_user(user):
        return base_qs

    department_ids = get_visible_active_memberships(user).values_list('department_id', flat=True)

    # Include system notifications, department notifications, and direct ticket notifications.
    return base_qs.filter(
        Q(ticket__isnull=True) |
        Q(ticket__assigned_department_id__in=department_ids) |
        Q(ticket__TICKET_CREATED=user) |
        Q(ticket__assigned_to=user) |
        Q(ticket__TICKET_CLOSED=user)
    )


def _redirect_to_safe_next(request, default_name, **kwargs):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return redirect(next_url)
    return redirect(default_name, **kwargs) if kwargs else redirect(default_name)

def landing_page(request):
    if request.user.is_authenticated:
        return redirect(_get_dashboard_redirect_url(request.user))
    return render(request, 'landing.html')

@login_required
def Basepage(request, dept_id=None):
    is_admin_user = _is_admin_user(request.user)
    user_memberships = DepartmentMember.objects.filter(
        user=request.user, is_active=True, department__is_active=True
    ).select_related('department').order_by('department__name')
    user_department_ids = list(user_memberships.values_list('department_id', flat=True))

    ticket_view = request.GET.get('view', '').strip().lower()
    is_created_view = (ticket_view == 'created')
    is_mine_only_filter = (
        request.GET.get('mine_only') in ['1', 'true', 'True', 'on']
        or request.GET.get('my_tickets') in ['1', 'true', 'True', 'on']
    )

    selected_department = None
    if not is_admin_user and user_department_ids and not is_created_view:
        if dept_id is not None:
            if dept_id not in user_department_ids:
                messages.error(request, "You do not have access to this department dashboard.")
                return redirect('base')
            selected_department = Department.objects.filter(id=dept_id).first()
    elif dept_id is not None and not is_created_view:
        selected_department = get_object_or_404(Department, id=dept_id)

    Ticketdatas = _exclude_inactive_department_tickets(TicketDetail.objects.all())
    if is_mine_only_filter:
        Ticketdatas = Ticketdatas.filter(TICKET_CREATED=request.user)
    elif is_created_view and not is_admin_user:
        Ticketdatas = Ticketdatas.filter(TICKET_CREATED=request.user)
    elif selected_department:
        Ticketdatas = Ticketdatas.filter(assigned_department=selected_department)
        if not is_admin_user:
            Ticketdatas = Ticketdatas.filter(
                ~Q(TICKET_CREATED=request.user)
            )
    elif not is_admin_user:
        department_ids = DepartmentMember.objects.filter(
            user=request.user, is_active=True
        ).values_list('department_id', flat=True)
        Ticketdatas = Ticketdatas.filter(
            Q(assigned_department_id__in=department_ids) &
            ~Q(TICKET_CREATED=request.user)
        ).distinct()

    if is_admin_user and not is_mine_only_filter:
        Ticketdatas = Ticketdatas.exclude(TICKET_CREATED=request.user)

    visible_tickets = Ticketdatas
    filter_form = TicketFilterForm(request.GET or None)
    if selected_department:
        filter_form.fields['department'].queryset = Department.objects.filter(id=selected_department.id)
        filter_form.fields['department'].initial = selected_department.id
    elif not is_admin_user:
        filter_form.fields['department'].queryset = Department.objects.filter(id__in=user_department_ids)

    if filter_form.is_valid():
        cd = filter_form.cleaned_data
        if cd.get('search'):
            Ticketdatas = Ticketdatas.filter(
                Q(TICKET_TITLE__icontains=cd['search']) |
                Q(TICKET_DESCRIPTION__icontains=cd['search'])
            )
        if cd.get('status'):
            Ticketdatas = Ticketdatas.filter(TICKET_STATUS=cd['status'])
        if cd.get('priority'):
            Ticketdatas = Ticketdatas.filter(priority=cd['priority'])
        if cd.get('category'):
            Ticketdatas = Ticketdatas.filter(category=cd['category'])
        if cd.get('department'):
            Ticketdatas = Ticketdatas.filter(assigned_department=cd['department'])
        if cd.get('my_tickets') or is_mine_only_filter:
            Ticketdatas = Ticketdatas.filter(TICKET_CREATED=request.user)

    paginator  = Paginator(Ticketdatas.order_by('-TICKET_CREATED_ON', '-id'), 20)
    page_obj   = paginator.get_page(request.GET.get('page'))

    if is_admin_user:
        ticket_ids = [ticket.id for ticket in page_obj.object_list]
        assignee_ids = {ticket.id: ticket.assigned_to_id for ticket in page_obj.object_list}
        preferred_rejections = {}
        fallback_rejections = {}
        if ticket_ids:
            rejected_histories = (
                TicketHistory.objects
                .filter(ticket_id__in=ticket_ids, action_type='REJECTED')
                .select_related('changed_by')
                .order_by('ticket_id', '-changed_at')
            )
            for history in rejected_histories:
                reason_text = (history.description or '').strip()
                if 'Reason:' in reason_text:
                    reason_text = reason_text.split('Reason:', 1)[1].strip()
                rejection_data = {
                    'reason': reason_text,
                    'rejected_by': history.changed_by.username if history.changed_by_id else 'Unknown',
                }
                current_assignee_id = assignee_ids.get(history.ticket_id)
                if current_assignee_id and history.changed_by_id == current_assignee_id:
                    if history.ticket_id not in fallback_rejections:
                        fallback_rejections[history.ticket_id] = rejection_data
                    continue
                if history.ticket_id not in preferred_rejections:
                    preferred_rejections[history.ticket_id] = rejection_data
        for ticket in page_obj.object_list:
            rejection_info = preferred_rejections.get(ticket.id) or fallback_rejections.get(ticket.id, {})
            ticket.latest_reject_reason = rejection_info.get('reason', '')
            ticket.latest_rejected_by = rejection_info.get('rejected_by', '')

    stats = {
        'total':       visible_tickets.count(),
        'open':        visible_tickets.filter(TICKET_STATUS='Open').count(),
        'in_progress': visible_tickets.filter(TICKET_STATUS='In Progress').count(),
        'closed':      visible_tickets.filter(TICKET_STATUS='Closed').count(),
        'resolved':    visible_tickets.filter(TICKET_STATUS='Resolved').count(),
        'my_tickets':    _exclude_inactive_department_tickets(
            TicketDetail.objects.filter(TICKET_CREATED=request.user)
        ).count(),
    }

    pagination_query = request.GET.copy()
    if 'page' in pagination_query:
        del pagination_query['page']

    department_dashboard_url = reverse('base')
    created_tickets_url = f"{reverse('base')}?view=created"

    return render(request, 'dashboard.html', {
        'Ticketdatas':   page_obj,
        'ticketdata':    page_obj,
        'filter_form': filter_form,
        'stats':       stats,
        'selected_department': selected_department,
        'is_created_view': is_created_view,
        'is_mine_only_filter': is_mine_only_filter,
        'is_admin_user': is_admin_user,
        'user_dashboard_departments': [m.department for m in user_memberships],
        'user_department_names': ", ".join([m.department.name for m in user_memberships]),
        'current_base_url': reverse('base_department', kwargs={'dept_id': selected_department.id}) if selected_department else reverse('base'),
        'department_dashboard_url': department_dashboard_url,
        'created_tickets_url': created_tickets_url,
        'pagination_query': pagination_query.urlencode(),
    })

@login_required
def TicketDetails(request):
    if _user_has_inactive_department_membership(request.user):
        messages.error(
            request,
            "You cannot create tickets while your department is inactive. Contact an administrator."
        )
        return redirect(_get_dashboard_redirect_url(request.user))

    if request.method == "POST":
        form = TicketCreateForm(request.POST, request.FILES)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.TICKET_CREATED = request.user

            prediction = {}
            used_ai_priority = True
            logger.info("AI priority prediction started for new ticket submit.")

            prediction = predict_ticket_priority_with_meta(
                title=ticket.TICKET_TITLE,
                description=ticket.TICKET_DESCRIPTION,
            )

            predicted_priority = (prediction or {}).get("priority")

            if predicted_priority:
                ticket.priority = predicted_priority
                ticket.ai_suggested_priority = predicted_priority
                logger.info("AI priority prediction applied: %s", predicted_priority)
            else:
                logger.warning(
                    "AI priority prediction returned no valid priority; falling back to default."
                )

            if not (ticket.priority or "").strip():
                ticket.priority = "MEDIUM"

            if not ticket.assigned_department:
                try:
                    predicted_department_name = predict_department(
                        title=ticket.TICKET_TITLE,
                        description=ticket.TICKET_DESCRIPTION,
                    )

                    normalized_prediction = (predicted_department_name or '').strip()
                    prediction_aliases = {
                        'human resources': 'HR',
                        'technical support': 'IT Support',
                        'customer service': 'Customer Support',
                    }
                    normalized_lookup = prediction_aliases.get(
                        normalized_prediction.lower(),
                        normalized_prediction,
                    )
                    predicted_department = Department.objects.filter(
                        Q(name__iexact=normalized_lookup) |
                        Q(code__iexact=normalized_lookup) |
                        Q(name__iexact=normalized_prediction) |
                        Q(code__iexact=normalized_prediction)
                    ).first()

                    if predicted_department and not predicted_department.is_active:
                        messages.error(request, f"{predicted_department.name} department is inactive. You cannot assign tickets to this department.")
                        return render(request, "TicketDetail.html", {"form": form})

                    if predicted_department and predicted_department.is_active:
                        ticket.assigned_department = predicted_department
                        ticket.assignment_type = "AUTO_ML"
                        ticket.assigned_at = timezone.now()

                        logger.info(
                            "ML department prediction applied: %s",
                            predicted_department_name,
                        )

                except Exception as e:
                    logger.warning("Department prediction failed: %s", str(e))

            if ticket.assigned_department:
                if not ticket.assignment_type:
                    ticket.assignment_type = "MANUAL"
                    ticket.assigned_by = request.user
                    ticket.assigned_at = timezone.now()

                if _is_creator_only_member_of_department(request.user, ticket.assigned_department):
                    messages.error(
                        request,
                        "Ticket cannot be submitted because AI routed it to your own department where you are the only active member."
                    )
                    return render(request, "TicketDetail.html", {"form": form})

            ticket.save()
            form.save_m2m()
            _auto_assign_single_member_department_ticket(ticket, changed_by=request.user)

            if used_ai_priority:
                _log_priority_prediction(ticket, prediction)

            TicketHistory.objects.create(
                ticket=ticket,
                changed_by=request.user,
                action_type="CREATED",
                description=f"Ticket created by {request.user.username}",
            )

            log_activity(
                request.user,
                "CREATED",
                f"Created ticket: {ticket.TICKET_TITLE}",
                ticket=ticket,
            )

            if ticket.assigned_department and ticket.assigned_department.is_active:
                notify_ticket_created(ticket)

                for member in DepartmentMember.objects.filter(
                    department=ticket.assigned_department,
                    is_active=True,
                    department__is_active=True,
                    user__is_active=True,
                ):
                    MyCart.objects.get_or_create(user=member.user, ticket=ticket)

            messages.success(request, "Ticket created successfully!")
            return redirect(_get_dashboard_redirect_url(request.user))

    else:
        form = TicketCreateForm()

    return render(request, "TicketDetail.html", {"form": form})

@login_required
def TicketInfo(request, pk):
    ticketinfos = get_object_or_404(TicketDetail, id=pk)
    if not _can_view_ticket(request.user, ticketinfos):
        is_same_department_user = (
            ticketinfos.assigned_department_id
            and DepartmentMember.objects.filter(
                user=request.user,
                department=ticketinfos.assigned_department,
                is_active=True,
            ).exists()
        )
        if (
            is_same_department_user
            and ticketinfos.assigned_to_id
            and ticketinfos.TICKET_CREATED_id != request.user.id
            and ticketinfos.assigned_to_id != request.user.id
        ):
            messages.error(request, "You have not permission to see this ticket.")
        else:
            messages.error(request, "You do not have permission to view this ticket.")
        return redirect('base')

    is_admin         = _is_admin_user(request.user)
    if request.method == 'POST' and is_admin and request.POST.get('action') == 'admin_reassign':
        return _handle_admin_reassign_ticket(request, ticketinfos)

    is_department_member = False
    is_senior_dept_member = False
    if ticketinfos.assigned_department:
        membership = DepartmentMember.objects.filter(
            user=request.user, department=ticketinfos.assigned_department, is_active=True
        ).first()
        if membership:
            is_department_member = True
            is_senior_dept_member = membership.role in ('SENIOR_MEMBER', 'LEAD', 'MANAGER')

    is_agent = is_admin or is_senior_dept_member
    _sync_mycart_for_user(request.user)
    can_work_on_ticket = _can_work_on_ticket(request.user, ticketinfos)
    can_close_ticket = _can_user_close_ticket(request.user, ticketinfos)
    can_edit_ticket = (
        ticketinfos.TICKET_STATUS == 'Open'
        and (ticketinfos.TICKET_CREATED_id == request.user.id or is_admin)
    )
    can_reject = (
        MyCart.objects.filter(user=request.user, ticket=ticketinfos).exists()
        and not _is_single_member_department_assignment(ticketinfos)
        and not _is_non_rejectable_assignment(request.user, ticketinfos)
    )
    comments_qs = UserComment.objects.filter(ticket=ticketinfos)
    if not is_admin:
        participant_ids = [ticketinfos.TICKET_CREATED_id]
        if ticketinfos.assigned_to_id:
            participant_ids.append(ticketinfos.assigned_to_id)
        comments_qs = comments_qs.filter(user_id__in=participant_ids)
    comments = comments_qs

    notification_id = request.GET.get('mark_read')
    if notification_id:
        try:
            Notification.objects.get(id=notification_id, user=request.user).mark_as_read()
        except Notification.DoesNotExist:
            pass

    creator_edit_only_mode = (
        request.GET.get('mine_only') in ['1', 'true', 'True', 'on']
        and ticketinfos.TICKET_CREATED_id == request.user.id
    )
    can_creator_decide_closed = (
        ticketinfos.TICKET_STATUS == 'Closed'
        and (ticketinfos.TICKET_CREATED_id == request.user.id or is_admin)
    )
    can_reopen_ticket = is_admin or not TicketHistory.objects.filter(
        ticket=ticketinfos,
        action_type='REOPENED',
        changed_by=request.user,
    ).exists()
    can_rate_resolved_ticket = (
        ticketinfos.TICKET_CREATED_id == request.user.id
        and ticketinfos.TICKET_STATUS == 'Resolved'
        and not _is_admin_user(request.user)
        and not hasattr(ticketinfos, 'rating')
    )
    normalized_description = " ".join((ticketinfos.TICKET_DESCRIPTION or "").split())
    can_view_admin_note_thread = bool(
        is_admin or ticketinfos.assigned_to_id == request.user.id
    )
    overdue_note_thread = None
    if can_view_admin_note_thread:
        overdue_note_paginator = Paginator(_overdue_note_thread_qs(ticketinfos), 4)
        requested_overdue_page = request.GET.get('overdue_page')
        overdue_page_number = requested_overdue_page or overdue_note_paginator.num_pages or 1
        overdue_note_thread = overdue_note_paginator.get_page(overdue_page_number)
    can_reply_admin_overdue_note = (
        can_view_admin_note_thread
        and not is_admin
        and ticketinfos.TICKET_STATUS not in ['Closed', 'Resolved']
    )
    latest_rejection = (
        TicketHistory.objects
        .filter(ticket=ticketinfos, action_type='REJECTED')
        .exclude(changed_by=ticketinfos.assigned_to)
        .select_related('changed_by')
        .order_by('-changed_at')
        .first()
    )
    if not latest_rejection:
        latest_rejection = (
            TicketHistory.objects
            .filter(ticket=ticketinfos, action_type='REJECTED')
            .select_related('changed_by')
            .order_by('-changed_at')
            .first()
        )
    latest_rejection_reason = ''
    latest_rejection_by = ''
    if latest_rejection:
        reason_text = (latest_rejection.description or '').strip()
        if 'Reason:' in reason_text:
            reason_text = reason_text.split('Reason:', 1)[1].strip()
        latest_rejection_reason = reason_text
        latest_rejection_by = latest_rejection.changed_by.username if latest_rejection.changed_by_id else ''
    admin_reassignment_options = _department_member_workload_rows(ticketinfos) if is_admin else []
    admin_reassignment_member_count = len(admin_reassignment_options)
    can_admin_reassign_ticket = bool(
        is_admin
        and ticketinfos.assigned_department_id
        and ticketinfos.assigned_to_id
        and ticketinfos.TICKET_STATUS not in ['Closed', 'Resolved']
        and any(not row['is_current_assignee'] for row in admin_reassignment_options)
    )

    return render(request, 'TicketInfo.html', {
        'ticketinfos':            ticketinfos,
        'normalized_description': normalized_description,
        'comments':             comments,
        'can_work_on_ticket':     can_work_on_ticket,
        'can_close_ticket':       can_close_ticket,
        'can_edit_ticket':        can_edit_ticket,
        'can_reject':           can_reject,
        'creator_edit_only_mode': creator_edit_only_mode,
        'can_creator_decide_closed': can_creator_decide_closed,
        'can_reopen_ticket': can_reopen_ticket,
        'can_rate_resolved_ticket': can_rate_resolved_ticket,
        'is_department_member': is_department_member,
        'is_agent':             is_agent,
        'is_admin':             is_admin,
        'latest_rejection_reason': latest_rejection_reason,
        'latest_rejection_by': latest_rejection_by,
        'can_view_admin_note_thread': can_view_admin_note_thread,
        'overdue_note_thread': overdue_note_thread,
        'can_reply_admin_overdue_note': can_reply_admin_overdue_note,
        'admin_reassignment_options': admin_reassignment_options,
        'admin_reassignment_member_count': admin_reassignment_member_count,
        'can_admin_reassign_ticket': can_admin_reassign_ticket,
    })


def _handle_admin_reassign_ticket(request, ticket):
    if ticket.TICKET_STATUS in ['Closed', 'Resolved']:
        messages.error(request, 'Only unresolved tickets can be reassigned.')
        return redirect('ticketinfo', pk=ticket.id)

    if not ticket.assigned_department_id or not getattr(ticket.assigned_department, 'is_active', False):
        messages.error(request, 'Only active department tickets can be reassigned.')
        return redirect('ticketinfo', pk=ticket.id)

    if not ticket.assigned_to_id:
        messages.error(request, 'This ticket must already be assigned before workload reassignment can be used.')
        return redirect('ticketinfo', pk=ticket.id)

    target_user_id = request.POST.get('assigned_to')
    if not target_user_id:
        messages.error(request, 'Select a department member for reassignment.')
        return redirect('ticketinfo', pk=ticket.id)

    eligible_memberships = _eligible_reassignment_memberships(ticket)
    target_membership = eligible_memberships.filter(user_id=target_user_id).first()
    if not target_membership:
        messages.error(request, 'Selected user is not an active member of this department.')
        return redirect('ticketinfo', pk=ticket.id)

    if target_membership.user_id == ticket.assigned_to_id:
        messages.error(request, 'That user is already assigned to this ticket.')
        return redirect('ticketinfo', pk=ticket.id)

    old_assignee = ticket.assigned_to
    ticket.assigned_to = target_membership.user
    ticket.TICKET_HOLDER = target_membership.user.username
    ticket.assignment_type = 'MANUAL'
    ticket.assigned_by = request.user
    ticket.assigned_at = timezone.now()
    ticket.save(update_fields=['assigned_to', 'TICKET_HOLDER', 'assignment_type', 'assigned_by', 'assigned_at'])

    dept_member_ids = eligible_memberships.values_list('user_id', flat=True)
    MyCart.objects.filter(ticket=ticket, user_id__in=dept_member_ids).exclude(user=target_membership.user).delete()
    MyCart.objects.get_or_create(user=target_membership.user, ticket=ticket)

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=request.user,
        action_type='ASSIGNED',
        field_name='assigned_to',
        old_value=old_assignee.username if old_assignee else 'Unassigned',
        new_value=target_membership.user.username,
        description=(
            f'Admin reassigned ticket to {target_membership.user.username} '
            f'based on workload review in {ticket.assigned_department.name}.'
        ),
    )
    create_notification(
        user=target_membership.user,
        notification_type='TICKET_ASSIGNED',
        title=f'Ticket reassigned #{ticket.id}',
        message=f'{request.user.username} reassigned "{ticket.TICKET_TITLE}" to you.',
        ticket=ticket,
        extra_data={
            'assigned_by': request.user.username,
            'department': ticket.assigned_department.name,
            'manual_workload_reassignment': True,
        },
    )

    messages.success(
        request,
        f'Ticket reassigned to {target_membership.user.get_full_name() or target_membership.user.username}.'
    )
    return redirect('ticketinfo', pk=ticket.id)


@login_required
def updateticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    is_admin_user = _is_admin_user(request.user)
    if is_admin_user:
        can_edit_ticket = ticket.TICKET_STATUS not in ['Closed', 'Resolved']
    else:
        can_edit_ticket = (
            ticket.TICKET_STATUS == 'Open'
            and ticket.TICKET_CREATED_id == request.user.id
        )
    if not can_edit_ticket:
        messages.error(request, 'Only the ticket creator can edit open tickets. Admin can edit unresolved tickets.')
        return redirect('ticketinfo', pk=pk)

    if not is_admin_user and not _can_work_on_ticket(request.user, ticket):
        messages.error(request, 'You do not have permission to edit this ticket.')
        return redirect('ticketinfo', pk=pk)

    if is_admin_user:
        old_department = ticket.assigned_department
        old_assignee = ticket.assigned_to
        if request.method == "POST":
            form = AdminTicketRoutingForm(request.POST, instance=ticket)
            if form.is_valid():
                changed_fields = [f for f in form.changed_data if f != 'extend_due_date']
                updated_ticket = form.save(commit=False)
                old_due_date = ticket.TICKET_DUE_DATE
                extended_due_date = form.cleaned_data.get('extend_due_date')
                if extended_due_date:
                    updated_ticket.TICKET_DUE_DATE = extended_due_date
                    changed_fields.append('TICKET_DUE_DATE')
                selected_department_id = request.POST.get('assigned_department')
                if (
                    selected_department_id
                    and old_department
                    and str(old_department.id) == str(selected_department_id)
                    and 'assigned_department' not in changed_fields
                    and 'priority' not in changed_fields
                    and 'TICKET_DUE_DATE' not in changed_fields
                ):
                    messages.error(request, 'Selected department is already assigned to this ticket.')
                    return render(request, 'Updateticket.html', {'form': form, 'ticket': ticket})

                if 'assigned_department' in changed_fields:
                    updated_ticket.assigned_to = None
                    updated_ticket.TICKET_HOLDER = ''
                if 'assigned_department' in changed_fields and updated_ticket.assigned_department:
                    updated_ticket.assignment_type = 'MANUAL'
                    updated_ticket.assigned_by = request.user
                    updated_ticket.assigned_at = timezone.now()
                    new_members = DepartmentMember.objects.filter(
                        department=updated_ticket.assigned_department,
                        is_active=True,
                        department__is_active=True,
                        user__is_active=True,
                    )
                    for member in new_members:
                        MyCart.objects.get_or_create(user=member.user, ticket=updated_ticket)
                        create_notification(
                            user=member.user, ticket=updated_ticket,
                            notification_type='TICKET_ASSIGNED',
                            title='Ticket reassigned',
                            message=f'Ticket "{updated_ticket.TICKET_TITLE}" reassigned to {updated_ticket.assigned_department.name}',
                        )

                updated_ticket.save()
                if (
                    'assigned_department' in changed_fields
                    and old_assignee
                ):
                    TicketHistory.objects.create(
                        ticket=updated_ticket,
                        changed_by=request.user,
                        action_type='ASSIGNED',
                        field_name='assigned_to',
                        old_value=old_assignee.username,
                        new_value='Unassigned',
                        description=(
                            f'Assignment cleared by {request.user.username} due to department reassignment.'
                        ),
                    )
                    create_notification(
                        user=old_assignee,
                        notification_type='TICKET_UPDATED',
                        title='Ticket reassigned to another department',
                        message=f'Ticket "{updated_ticket.TICKET_TITLE}" was moved to {updated_ticket.assigned_department.name if updated_ticket.assigned_department else "another department"} and unassigned from you.',
                        ticket=updated_ticket,
                        extra_data={
                            'updated_by': request.user.username,
                            'old_department': old_department.name if old_department else '',
                            'new_department': updated_ticket.assigned_department.name if updated_ticket.assigned_department else '',
                        },
                    )
                _auto_assign_single_member_department_ticket(updated_ticket, changed_by=request.user)
                form.save_m2m()

                if 'assigned_department' in changed_fields:
                    if old_department and old_department != updated_ticket.assigned_department:
                        old_member_ids = DepartmentMember.objects.filter(
                            department=old_department, is_active=True
                        ).values_list('user_id', flat=True)
                        if updated_ticket.assigned_department:
                            new_member_ids = DepartmentMember.objects.filter(
                                department=updated_ticket.assigned_department, is_active=True
                            ).values_list('user_id', flat=True)
                            MyCart.objects.filter(
                                ticket=updated_ticket, user_id__in=old_member_ids
                            ).exclude(user_id__in=new_member_ids).delete()
                        else:
                            MyCart.objects.filter(ticket=updated_ticket, user_id__in=old_member_ids).delete()

                for field in changed_fields:
                    action_type = 'PRIORITY_CHANGED' if field == 'priority' else 'UPDATED'
                    old_value = form.initial.get(field, '')
                    new_value = form.cleaned_data.get(field, '')
                    description = f'{field} changed by {request.user.username}'
                    if field == 'TICKET_DUE_DATE':
                        old_value = old_due_date
                        new_value = updated_ticket.TICKET_DUE_DATE
                        description = f'Due date extended by {request.user.username}'
                    TicketHistory.objects.create(
                        ticket=updated_ticket, changed_by=request.user,
                        action_type=action_type, field_name=field,
                        old_value=str(old_value),
                        new_value=str(new_value),
                        description=description,
                    )

                if 'TICKET_DUE_DATE' in changed_fields:
                    notify_ticket_due_date_extended(
                        updated_ticket,
                        request.user,
                        old_due_date,
                        updated_ticket.TICKET_DUE_DATE,
                    )
                other_changes = [f for f in changed_fields if f != 'TICKET_DUE_DATE']
                if other_changes:
                    notify_ticket_updated(updated_ticket, request.user, changes=changed_fields)
                log_activity(request.user, 'UPDATED', f'Updated ticket routing: {updated_ticket.TICKET_TITLE}', ticket=updated_ticket)
                if changed_fields == ['TICKET_DUE_DATE']:
                    messages.success(request, 'Ticket due date extended successfully.')
                elif 'TICKET_DUE_DATE' in changed_fields:
                    messages.success(request, 'Ticket updated successfully, including due date extension.')
                else:
                    messages.success(request, 'Ticket priority/department updated successfully.')
                return redirect('ticketinfo', pk=pk)
        else:
            form = AdminTicketRoutingForm(instance=ticket)

        return render(request, 'Updateticket.html', {'form': form, 'ticket': ticket})

    is_creator_edit = ticket.TICKET_CREATED_id == request.user.id and not is_admin_user
    if is_creator_edit:
        old_department = ticket.assigned_department
        form = TicketDetailForm(request.POST or None, request.FILES or None, instance=ticket)
        if request.method == "POST" and form.is_valid():
            changed_fields = form.changed_data
            updated_ticket = form.save(commit=False)
            updated_ticket.TICKET_CREATED = ticket.TICKET_CREATED
            if updated_ticket.TICKET_DESCRIPTION:
                updated_ticket.TICKET_DESCRIPTION = " ".join(updated_ticket.TICKET_DESCRIPTION.split())

            if 'assigned_department' in changed_fields and updated_ticket.assigned_department:
                updated_ticket.assignment_type = 'MANUAL'
                updated_ticket.assigned_by = request.user
                updated_ticket.assigned_at = timezone.now()
                new_members = DepartmentMember.objects.filter(
                    department=updated_ticket.assigned_department,
                    is_active=True,
                    department__is_active=True,
                    user__is_active=True,
                )
                for member in new_members:
                    MyCart.objects.get_or_create(user=member.user, ticket=updated_ticket)
                    create_notification(
                        user=member.user, ticket=updated_ticket,
                        notification_type='TICKET_ASSIGNED',
                        title='Ticket reassigned',
                        message=f'Ticket "{updated_ticket.TICKET_TITLE}" reassigned to {updated_ticket.assigned_department.name}',
                    )

            updated_ticket.save()
            _auto_assign_single_member_department_ticket(updated_ticket, changed_by=request.user)
            form.save_m2m()
            if 'priority' in changed_fields:
                _record_priority_feedback(updated_ticket, form.cleaned_data.get('priority'), request.user)

            if 'assigned_department' in changed_fields and old_department and old_department != updated_ticket.assigned_department:
                old_member_ids = DepartmentMember.objects.filter(
                    department=old_department, is_active=True
                ).values_list('user_id', flat=True)
                if updated_ticket.assigned_department:
                    new_member_ids = DepartmentMember.objects.filter(
                        department=updated_ticket.assigned_department, is_active=True
                    ).values_list('user_id', flat=True)
                    MyCart.objects.filter(
                        ticket=updated_ticket, user_id__in=old_member_ids
                    ).exclude(user_id__in=new_member_ids).delete()
                else:
                    MyCart.objects.filter(ticket=updated_ticket, user_id__in=old_member_ids).delete()

            for field in changed_fields:
                action_type = 'PRIORITY_CHANGED' if field == 'priority' else 'UPDATED'
                TicketHistory.objects.create(
                    ticket=updated_ticket, changed_by=request.user,
                    action_type=action_type, field_name=field,
                    old_value=str(form.initial.get(field, '')),
                    new_value=str(form.cleaned_data.get(field, '')),
                    description=f'{field} changed by {request.user.username}',
                )

            notify_ticket_updated(updated_ticket, request.user, changes=changed_fields)
            log_activity(request.user, 'UPDATED', f'Updated ticket: {updated_ticket.TICKET_TITLE}', ticket=updated_ticket)
            messages.success(request, 'Ticket updated successfully!')
            if request.GET.get('mine_only') in ['1', 'true', 'True', 'on']:
                return redirect(f"{reverse('ticketinfo', kwargs={'pk': pk})}?mine_only=1")
            return redirect('ticketinfo', pk=pk)

        return render(request, 'TicketDetail.html', {
            'form': form,
            'ticket': ticket,
            'edit_mode': True,
            'mine_only_mode': request.GET.get('mine_only') in ['1', 'true', 'True', 'on'],
        })

    if request.method == "POST":
        old_department = ticket.assigned_department
        form = TicketUpdateForm(request.POST, instance=ticket)
        if form.is_valid():
            changed_fields = form.changed_data
            updated_ticket   = form.save(commit=False)
            if updated_ticket.TICKET_DESCRIPTION:
                updated_ticket.TICKET_DESCRIPTION = " ".join(updated_ticket.TICKET_DESCRIPTION.split())

            if 'assigned_department' in changed_fields and updated_ticket.assigned_department:
                updated_ticket.assignment_type = 'MANUAL'
                updated_ticket.assigned_by     = request.user
                updated_ticket.assigned_at     = timezone.now()
                new_members = DepartmentMember.objects.filter(
                    department=updated_ticket.assigned_department,
                    is_active=True,
                    department__is_active=True,
                    user__is_active=True,
                )
                for member in new_members:
                    MyCart.objects.get_or_create(user=member.user, ticket=updated_ticket)
                    create_notification(
                        user=member.user, ticket=updated_ticket,
                        notification_type='TICKET_ASSIGNED',
                        title='Ticket reassigned',
                        message=f'Ticket "{updated_ticket.TICKET_TITLE}" reassigned to {updated_ticket.assigned_department.name}',
                    )
            if 'assigned_to' in changed_fields and updated_ticket.assigned_to:
                updated_ticket.assignment_type = 'MANUAL'
                updated_ticket.assigned_by = request.user
                updated_ticket.assigned_at = timezone.now()

            updated_ticket.save()
            _auto_assign_single_member_department_ticket(updated_ticket, changed_by=request.user)
            form.save_m2m()
            if 'priority' in changed_fields:
                _record_priority_feedback(updated_ticket, form.cleaned_data.get('priority'), request.user)

            if 'assigned_to' in changed_fields and updated_ticket.assigned_to:
                MyCart.objects.get_or_create(user=updated_ticket.assigned_to, ticket=updated_ticket)
                MyCart.objects.filter(ticket=updated_ticket).exclude(user=updated_ticket.assigned_to).delete()

            if 'assigned_department' in changed_fields:
                if old_department and old_department != updated_ticket.assigned_department:
                    old_member_ids = DepartmentMember.objects.filter(
                        department=old_department, is_active=True
                    ).values_list('user_id', flat=True)
                    if updated_ticket.assigned_department:
                        new_member_ids = DepartmentMember.objects.filter(
                            department=updated_ticket.assigned_department, is_active=True
                        ).values_list('user_id', flat=True)
                        MyCart.objects.filter(
                            ticket=updated_ticket, user_id__in=old_member_ids
                        ).exclude(user_id__in=new_member_ids).delete()
                    else:
                        MyCart.objects.filter(ticket=updated_ticket, user_id__in=old_member_ids).delete()

            for field in changed_fields:
                action_type = 'STATUS_CHANGED' if field == 'TICKET_STATUS' else \
                              'PRIORITY_CHANGED' if field == 'priority' else 'UPDATED'
                TicketHistory.objects.create(
                    ticket=updated_ticket, changed_by=request.user,
                    action_type=action_type, field_name=field,
                    old_value=str(form.initial.get(field, '')),
                    new_value=str(form.cleaned_data.get(field, '')),
                    description=f'{field} changed by {request.user.username}',
                )

            notify_ticket_updated(ticket, request.user, changes=changed_fields)
            log_activity(request.user, 'UPDATED', f'Updated ticket: {updated_ticket.TICKET_TITLE}', ticket=updated_ticket)
            messages.success(request, 'Ticket updated successfully!')
            return redirect('ticketinfo', pk=pk)
    else:
        form = TicketUpdateForm(instance=ticket)
    return render(request, 'Updateticket.html', {'form': form, 'ticket': ticket})


@admin_required
def admin_reassign_ticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if request.method != 'POST':
        return redirect('ticketinfo', pk=pk)
    return _handle_admin_reassign_ticket(request, ticket)


@login_required
def deleteticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('ticketinfo', pk=pk)
    inactive_ticket_locked = (
        not _is_admin_user(request.user)
        and ticket.__class__.objects.filter(id=ticket.id).filter(_inactive_department_ticket_q()).exists()
    )
    if inactive_ticket_locked:
        messages.error(request, 'You do not have permission to delete this ticket.')
        return redirect('base')
    can_delete_by_role = user_has_department_permission(
        request.user, ticket.assigned_department, 'can_delete_tickets'
    )
    if ticket.TICKET_CREATED != request.user and not request.user.is_superuser and not can_delete_by_role:
        messages.error(request, 'You do not have permission to delete this ticket.')
        return redirect('base')

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=request.user,
        action_type='DELETED',
        description=f'Ticket deleted by {request.user.username}',
    )

    log_activity(
        request.user,
        'DELETED',
        f'Deleted ticket: {ticket.TICKET_TITLE}'
    )

    ticket.delete()
    messages.success(request, 'Ticket deleted successfully!')
    return redirect(_get_dashboard_redirect_url(request.user))


@admin_required
def bulk_delete_tickets(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('base')

    raw_ids = request.POST.get('ticket_ids', '')
    ticket_ids = []
    for value in raw_ids.split(','):
        value = value.strip()
        if value.isdigit():
            ticket_ids.append(int(value))

    if not ticket_ids:
        messages.error(request, 'No tickets selected for deletion.')
        next_url = request.POST.get('next')
        if next_url and url_has_allowed_host_and_scheme(
            next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
        ):
            return redirect(next_url)
        return redirect('base')

    tickets = list(TicketDetail.objects.filter(id__in=ticket_ids))
    if not tickets:
        messages.error(request, 'Selected tickets were not found.')
        next_url = request.POST.get('next')
        if next_url and url_has_allowed_host_and_scheme(
            next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
        ):
            return redirect(next_url)
        return redirect('base')

    for ticket in tickets:
        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=request.user,
            action_type='DELETED',
            description=f'Ticket deleted by {request.user.username} (bulk delete)',
        )
        log_activity(
            request.user,
            'DELETED',
            f'Deleted ticket: {ticket.TICKET_TITLE}'
        )

    TicketDetail.objects.filter(id__in=[t.id for t in tickets]).delete()
    messages.success(request, f'Deleted {len(tickets)} ticket(s) successfully.')

    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return redirect(next_url)
    return redirect('base')

@login_required
def RemoveTicket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if not _can_view_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('ticketinfo', pk=pk)
    if _is_single_member_department_assignment(ticket):
        messages.error(request, "You cannot reject this ticket because you are the only active member in this department.")
        return redirect('ticketinfo', pk=pk)
    if _is_non_rejectable_assignment(request.user, ticket):
        messages.error(request, "This assignment cannot be rejected.")
        return redirect('ticketinfo', pk=pk)
    if not MyCart.objects.filter(ticket=ticket, user=request.user).exists() and ticket.assigned_to_id != request.user.id:
        messages.error(request, "Only the current assignee can reject this ticket.")
        return redirect('ticketinfo', pk=pk)

    if request.method != 'POST':
        return render(request, 'reject_ticket.html', {'ticket': ticket})

    rejected = MyCart.objects.filter(ticket=ticket, user=request.user).exists()
    reason = request.POST.get('reject_reason', '').strip()
    if not reason:
        messages.error(request, "Rejection reason is required.")
        return redirect('ticketinfo', pk=pk)
    MyCart.objects.filter(ticket=ticket, user=request.user).delete()

    auto_assignee = None
    if ticket.assigned_to_id == request.user.id:
        old_status = ticket.TICKET_STATUS
        ticket.assigned_to = None
        ticket.TICKET_HOLDER = ''
        if ticket.TICKET_STATUS in ['In Progress', 'Reopen']:
            ticket.TICKET_STATUS = 'Open'
        ticket.save(update_fields=['assigned_to', 'TICKET_HOLDER', 'TICKET_STATUS'])
        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=request.user,
            action_type='STATUS_CHANGED',
            old_value=old_status,
            new_value=ticket.TICKET_STATUS,
            description=f'Ticket released by {request.user.username}'
                        + (f'. Reason: {reason}' if reason else ''),
        )
        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=request.user,
            action_type='REJECTED',
            description=f'Ticket rejected by {request.user.username}'
                        + (f'. Reason: {reason}' if reason else ''),
        )
        auto_assignee = _auto_assign_on_department_rejection(ticket, request.user)
    elif rejected:
        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=request.user,
            action_type='REJECTED',
            description=f'Ticket rejected from queue by {request.user.username}'
                        + (f'. Reason: {reason}' if reason else ''),
        )
        auto_assignee = _auto_assign_on_department_rejection(ticket, request.user)
    if auto_assignee:
        messages.success(
            request,
            f"Ticket rejected and auto-assigned to {auto_assignee.get_full_name() or auto_assignee.username}."
        )
    else:
        messages.success(request, "Ticket rejected and removed from your queue.")
    return redirect('mycart')

@login_required
def CloseTicket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if _is_close_locked_by_rejection(request.user, ticket):
        messages.error(request, "You have already rejected this ticket and cannot close it now.")
        return redirect('ticketinfo', pk=pk)
    if _is_admin_user(request.user):
        messages.error(request, "Admins cannot close tickets. You can delete the ticket if needed.")
        return redirect('ticketinfo', pk=pk)
    if not _can_user_close_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('ticketinfo', pk=pk)
    # Capture old status BEFORE changing it
    old_status = ticket.TICKET_STATUS
    ticket.assigned_to = request.user
    ticket.TICKET_STATUS   = 'Closed'
    ticket.TICKET_CLOSED   = request.user
    ticket.TICKET_CLOSED_ON = timezone.now()
    ticket.save()
    TicketHistory.objects.create(
        ticket=ticket, changed_by=request.user,
        action_type='CLOSED',
        description=f'Ticket closed by {request.user.username}',
        old_value=old_status, new_value='Closed',
    )
    log_activity(request.user, 'CLOSED', f'Closed ticket: {ticket.TICKET_TITLE}', ticket=ticket)
    notify_ticket_closed(ticket, request.user)
    if ticket.assigned_department_id:
        dept_member_ids = DepartmentMember.objects.filter(
            department=ticket.assigned_department,
            is_active=True
        ).values_list('user_id', flat=True)
        MyCart.objects.filter(ticket=ticket, user_id__in=dept_member_ids).delete()
    else:
        MyCart.objects.filter(ticket=ticket).delete()
    messages.success(request, "Ticket closed successfully.")
    return redirect(_get_dashboard_redirect_url(request.user))

@login_required
def reopenticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    is_admin = _is_admin_user(request.user)
    
    if ticket.TICKET_STATUS not in ['Closed', 'Resolved', 'Reopen']:
        messages.error(request, "This ticket cannot be reopened as it has not been closed.")
        return redirect('ticketinfo', pk=pk)
    
    if (
        not is_admin
        and TicketHistory.objects.filter(
            ticket=ticket,
            action_type='REOPENED',
            changed_by=request.user,
        ).exists()
    ):
        messages.error(request, "You can reopen a ticket only once. Please create a new ticket.")
        return redirect('ticketinfo', pk=pk)
    if ticket.TICKET_CREATED_id != request.user.id and not is_admin:
        messages.error(request, "Only the ticket creator can reopen this ticket.")
        return redirect('ticketinfo', pk=pk)
    if not is_admin and not _can_work_on_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('ticketinfo', pk=pk)
    holder = ticket.TICKET_CLOSED

    if not holder:
        messages.error(request, "Cannot reopen: no closer recorded.")
        return redirect('base')

    old_status = ticket.TICKET_STATUS
    ticket.TICKET_STATUS = 'Reopen'
    ticket.TICKET_CLOSED_ON = None
    ticket.resolved_at = None
    ticket.assigned_to = holder
    ticket.TICKET_HOLDER = holder.username
    ticket.save(update_fields=[
        'TICKET_STATUS',
        'TICKET_CLOSED_ON',
        'resolved_at',
        'assigned_to',
        'TICKET_HOLDER',
    ])

    MyCart.objects.get_or_create(user=holder, ticket=ticket)

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=request.user,
        action_type='REOPENED',
        description=f'Ticket reopened by {request.user.username}',
        old_value=old_status,
        new_value='Reopen',
    )

    log_activity(
        request.user,
        'REOPENED',
        f'Reopened ticket: {ticket.TICKET_TITLE}',
        ticket=ticket
    )

    notify_ticket_reopened(ticket, request.user)

    messages.success(request, "Ticket reopened successfully.")
    return redirect(_get_dashboard_redirect_url(request.user))

@login_required
def resolvedticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if ticket.TICKET_STATUS == 'Resolved':
        messages.info(request, "This ticket is already resolved.")
        return redirect('ticketinfo', pk=pk)
    is_admin = _is_admin_user(request.user)
    if ticket.TICKET_STATUS == 'Closed' and ticket.TICKET_CREATED_id != request.user.id and not is_admin:
        messages.error(request, "Only the ticket creator can resolve a closed ticket.")
        return redirect('ticketinfo', pk=pk)
    if not is_admin and not _can_work_on_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('ticketinfo', pk=pk)
    old_status = ticket.TICKET_STATUS

    ticket.TICKET_STATUS = 'Resolved'
    ticket.resolved_at = timezone.now()
    ticket.save()

    notify_ticket_resolved(ticket, request.user)

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=request.user,
        action_type='STATUS_CHANGED',
        description=f'Ticket resolved by {request.user.username}',
        old_value=old_status,
        new_value='Resolved',
    )

    log_activity(
        request.user,
        'RESOLVED',
        f'Resolved ticket: {ticket.TICKET_TITLE}',
        ticket=ticket
    )

    messages.success(request, "Ticket resolved successfully.")
    return redirect('ticketinfo', pk=pk)

def LoginView(request):
    if request.user.is_authenticated:
        return redirect(_get_dashboard_redirect_url(request.user))

    role = LoginRoleAuthorization.normalize_mode(request.GET.get('role'))

    if request.method == "POST":
        role = LoginRoleAuthorization.normalize_mode(request.POST.get('login_as', role))
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
            )
            if user:
                _ensure_userprofile_and_permissions(user)

                                                                                
                access_error = LoginRoleAuthorization.account_access_error(role, user)
                if access_error:
                    messages.error(request, access_error)
                    return render(request, 'Login.html', {
                        'form': form,
                        'role': role,
                        'allow_register': LoginRoleAuthorization.can_register(role),
                    })

                login(request, user)
                if not form.cleaned_data.get('remember_me'):
                    request.session.set_expiry(0)

                if role == LoginRoleAuthorization.ADMIN:
                    messages.success(request, f"Welcome, {user.username}! You are logged in as Admin.")
                else:
                    messages.success(request, f"Welcome back, {user.get_full_name() or user.username}!")

                                                                              
                return redirect(LoginRoleAuthorization.success_redirect(
                    role, user, _get_dashboard_redirect_url
                ))
            else:
                messages.error(request, "Invalid username or password.")
    else:
        form = LoginForm(initial={'login_as': role})

    return render(request, 'Login.html', {
        'form': form,
        'role': role,
        'allow_register': LoginRoleAuthorization.can_register(role),
    })


def LogoutView(request):
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect('landing')


def RegisterView(request):
    if request.method == "POST":
        register_form = RegisterForm(request.POST)
        profile_form  = UserProfileForm(
            request.POST,
            require_phone=True,
            include_profile_image=False,
        )
        if register_form.is_valid() and profile_form.is_valid():
            user         = register_form.save()
            profile      = profile_form.save(commit=False)
            profile.user = user
            profile.phone = profile_form.get_full_phone()
            profile.save()
            messages.success(request, "Registered successfully! Please sign in.")
            return redirect('login')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        register_form = RegisterForm()
        profile_form  = UserProfileForm(require_phone=True, include_profile_image=False)
    return render(request, 'Register.html', {
        'Register_form': register_form,
        'UserProfile_form': profile_form,
    })

@login_required
def Change_Password(request):
    if request.method == 'POST':
        fm = PasswordChangeForm(user=request.user, data=request.POST)
        if fm.is_valid():
            fm.save()
            update_session_auth_hash(request, fm.user)
            log_activity(request.user, 'UPDATED', 'Changed account password')
            messages.success(request, 'Password changed successfully.')
            return redirect('login')
    else:
        fm = PasswordChangeForm(user=request.user)
    return render(request, 'Change_Password.html', {'fm': fm})

@login_required
def User_Profile(request):
    user = request.user
    profile_data = UserProfile.objects.filter(user=user)
    resolved_count = TicketDetail.objects.filter(
        assigned_to=user, TICKET_STATUS__in=['Closed', 'Resolved']
    ).count()
    avg_rating = TicketRating.objects.filter(
        ticket__assigned_to=user
    ).aggregate(avg=Avg('rating'))['avg']
    return render(request, 'Userprofile.html', {
        'ProfileDatas':    profile_data,
        'resolved_count':  resolved_count,
        'avg_rating':      round(avg_rating, 1) if avg_rating else None,
    })


@login_required
def update_profile(request, pk):
    profile = get_object_or_404(UserProfile, id=pk)
    if profile.user != request.user and not request.user.is_superuser:
        messages.error(request, "You do not have permission to edit this profile.")
        return redirect('profile')
    form = UserProfileForm(request.POST or None, request.FILES or None, instance=profile)
    if form.is_valid():
        form.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('profile')
    return render(request, 'Update_Profile.html', {'form': form})

@login_required
def MyCarts(request):
    _sync_mycart_for_user(request.user)
    user_memberships = DepartmentMember.objects.filter(
        user=request.user, is_active=True, department__is_active=True
    ).select_related('department').order_by('department__name')
    user_department_ids = list(user_memberships.values_list('department_id', flat=True))

    base_carts = MyCart.objects.filter(user=request.user).select_related(
        'ticket', 'ticket__TICKET_CREATED', 'ticket__assigned_department', 'ticket__category'
    ).exclude(_inactive_department_ticket_q('ticket__')).order_by('-accepted_at')
    total_assigned_count = base_carts.count()
    carts = base_carts

    department_filter = request.GET.get('department', '').strip()
    if department_filter:
        try:
            dept_id = int(department_filter)
        except (TypeError, ValueError):
            dept_id = None
        if dept_id and dept_id in user_department_ids:
            carts = carts.filter(ticket__assigned_department_id=dept_id)

    sort = (request.GET.get('sort', 'all') or 'all').strip().lower()
    if sort not in ['all', 'low', 'medium', 'high', 'urgent', 'overdue']:
        sort = 'all'
    today = timezone.localdate()
    priority_filter_map = {
        'low': 'LOW',
        'medium': 'MEDIUM',
        'high': 'HIGH',
        'urgent': 'URGENT',
    }
    if sort in priority_filter_map:
        priority_code = priority_filter_map[sort]
        carts = carts.filter(
            Q(ticket__priority=priority_code) | Q(ticket__ai_suggested_priority=priority_code)
        )
    elif sort == 'overdue':
        carts = carts.filter(ticket__TICKET_DUE_DATE__lt=today).exclude(
            ticket__TICKET_STATUS__in=['Closed', 'Resolved']
        )

    comments = UserComment.objects.filter(user=request.user)
    cart_ticket_ids = list(carts.values_list('ticket_id', flat=True))
    non_rejectable_ticket_ids = set()
    reopened_ticket_ids = set(
        carts.filter(ticket__TICKET_STATUS='Reopen').values_list('ticket_id', flat=True)
    )
    non_rejectable_ticket_ids.update(reopened_ticket_ids)
    admin_assigned_ticket_ids = set(
        carts.filter(ticket__assigned_by__is_superuser=True).values_list('ticket_id', flat=True)
    )
    non_rejectable_ticket_ids.update(admin_assigned_ticket_ids)
    single_member_department_ticket_ids = {
        cart.ticket_id for cart in carts if _is_single_member_department_assignment(cart.ticket)
    }
    non_rejectable_ticket_ids.update(single_member_department_ticket_ids)
    if cart_ticket_ids:
        latest_assigned_history = (
            TicketHistory.objects.filter(
                ticket_id__in=cart_ticket_ids,
                action_type='ASSIGNED',
            )
            .order_by('ticket_id', '-changed_at')
        )
        latest_reopened_history = (
            TicketHistory.objects.filter(
                ticket_id__in=cart_ticket_ids,
                action_type='REOPENED',
            )
            .order_by('ticket_id', '-changed_at')
        )

        latest_assigned_by_ticket = {}
        for row in latest_assigned_history:
            if row.ticket_id not in latest_assigned_by_ticket:
                latest_assigned_by_ticket[row.ticket_id] = row

        latest_reopened_by_ticket = {}
        for row in latest_reopened_history:
            if row.ticket_id not in latest_reopened_by_ticket:
                latest_reopened_by_ticket[row.ticket_id] = row

        for ticket_id, assigned_row in latest_assigned_by_ticket.items():
            reopened_row = latest_reopened_by_ticket.get(ticket_id)
            was_reopened_after_assignment = (
                reopened_row is not None and reopened_row.changed_at > assigned_row.changed_at
            )
            if (
                assigned_row.new_value == request.user.username
                and assigned_row.description
                and 'Auto-assigned to' in assigned_row.description
                and not was_reopened_after_assignment
            ):
                non_rejectable_ticket_ids.add(ticket_id)

    stats = {
        'assigned': carts.count(),
        'in_progress': carts.filter(ticket__TICKET_STATUS='In Progress').count(),
        'overdue': carts.filter(ticket__TICKET_DUE_DATE__lt=today).exclude(
            ticket__TICKET_STATUS__in=['Closed', 'Resolved']
        ).count(),
    }
    closable_ticket_ids = {
        cart.ticket_id
        for cart in carts
        if _can_user_close_ticket(request.user, cart.ticket)
    }
    return render(request, 'Mycart.html', {
        'Carts': carts,
        'comments': comments,
        'stats': stats,
        'sort': sort,
        'total_assigned_count': total_assigned_count,
        'has_active_ticket_filter': (sort != 'all') or bool(department_filter),
        'user_departments': [m.department for m in user_memberships],
        'selected_department_filter': department_filter,
        'non_rejectable_ticket_ids': non_rejectable_ticket_ids,
        'closable_ticket_ids': closable_ticket_ids,
    })

@login_required
def activity_log(request):
    logs = ActivityLog.objects.filter(user=request.user).select_related('ticket')

    action_filter = request.GET.get('action', '')
    if action_filter:
        logs = logs.filter(action=action_filter)

    search = request.GET.get('search', '').strip()
    if search:
        logs = logs.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(ticket__TICKET_TITLE__icontains=search)
        )

    paginator = Paginator(logs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    stats = {
        'total':    ActivityLog.objects.filter(user=request.user).count(),
        'resolved': ActivityLog.objects.filter(user=request.user, action='RESOLVED').count(),
        'comments': ActivityLog.objects.filter(user=request.user, action='COMMENTED').count(),
    }

    return render(request, 'activity_log.html', {
        'activities':    page_obj,
        'stats':         stats,
        'action_filter': action_filter,
        'search':        search,
    })

@login_required
def resolved_history(request):
    resolved = _exclude_inactive_department_tickets(TicketDetail.objects.filter(
        TICKET_STATUS__in=['Resolved', 'Closed'],
    ))

    q = request.GET.get('q', '').strip()
    search_by = (request.GET.get('search_by_btn') or request.GET.get('search_by', 'all')).strip().lower()
    if search_by not in {'all', 'id', 'title', 'description', 'name', 'department'}:
        search_by = 'all'
    if q:
        if search_by == 'id':
            if q.isdigit():
                resolved = resolved.filter(id=int(q))
            else:
                resolved = resolved.none()
        elif search_by == 'title':
            resolved = resolved.filter(TICKET_TITLE__icontains=q)
        elif search_by == 'description':
            resolved = resolved.filter(TICKET_DESCRIPTION__icontains=q)
        elif search_by == 'name':
            resolved = resolved.filter(
                Q(TICKET_CREATED__username__icontains=q) |
                Q(TICKET_CREATED__first_name__icontains=q) |
                Q(TICKET_CREATED__last_name__icontains=q) |
                Q(assigned_to__username__icontains=q) |
                Q(assigned_to__first_name__icontains=q) |
                Q(assigned_to__last_name__icontains=q)
            )
        elif search_by == 'department':
            resolved = resolved.filter(assigned_department__name__icontains=q)
        else:
            resolved = resolved.filter(
                Q(TICKET_TITLE__icontains=q) |
                Q(TICKET_DESCRIPTION__icontains=q) |
                Q(TICKET_CREATED__username__icontains=q) |
                Q(TICKET_CREATED__first_name__icontains=q) |
                Q(TICKET_CREATED__last_name__icontains=q) |
                Q(assigned_to__username__icontains=q) |
                Q(assigned_to__first_name__icontains=q) |
                Q(assigned_to__last_name__icontains=q) |
                Q(assigned_department__name__icontains=q)
            )

    resolved = resolved.select_related(
        'category', 'assigned_to', 'TICKET_CREATED', 'assigned_department'
    ).order_by('-TICKET_CLOSED_ON')

    full_qs = _exclude_inactive_department_tickets(
        TicketDetail.objects.filter(TICKET_STATUS__in=['Resolved', 'Closed'])
    )
    stats = {
        'total_resolved': full_qs.count(),
        'resolved':       full_qs.filter(TICKET_STATUS='Resolved').count(),
        'closed':         full_qs.filter(TICKET_STATUS='Closed').count(),
        'avg_rating':     TicketRating.objects.filter(
            ticket__in=full_qs
        ).aggregate(avg=Avg('rating'))['avg'],
    }

    paginator = Paginator(resolved, 5)
    page_obj  = paginator.get_page(request.GET.get('page'))

    return render(request, 'resolved_history.html', {
        'resolved_tickets': page_obj,
        'stats': stats,
        'search_by': search_by,
    })

@admin_required
def account_settings(request):
    form = AccountSettingsForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        target_user = form.cleaned_data['target_user']
        action      = form.cleaned_data['action']

        if action == 'department':
            dept = form.cleaned_data.get('department')
            if dept:
                DepartmentMember.objects.get_or_create(
                    user=target_user, department=dept,
                    defaults={'role': 'MEMBER', 'added_by': request.user}
                )
                log_activity(
                    request.user, 'UPDATED',
                    f'Added {target_user.username} to {dept.name}',
                )
                messages.success(request, f'{target_user.username} added to {dept.name}.')

        elif action == 'toggle_status':
            target_user.is_active = not target_user.is_active
            target_user.save()
            status = 'activated' if target_user.is_active else 'deactivated'
            log_activity(
                request.user, 'UPDATED',
                f'Account {status}: {target_user.username}',
            )
            messages.success(request, f'{target_user.username} account {status}.')

        return redirect('account_settings')

    stats = {
        'total_users':    User.objects.count(),
        'superusers':     User.objects.filter(is_superuser=True).count(),
        'departments':    Department.objects.filter(is_active=True).count(),
    }

    return render(request, 'account_settings.html', {
        'form':  form,
        'stats': stats,
        'total_users': stats['total_users'],
        'superusers': stats['superusers'],
        'active_admins': stats['superusers'],
        'dept_count': stats['departments'],
    })

@admin_required
def dashboard_pie(request):
    statuses = ['Open', 'In Progress', 'Reopen', 'Resolved', 'Closed']
    counts = [TicketDetail.objects.filter(TICKET_STATUS=s).count() for s in statuses]
    colors = ['#3B82F6', '#F59E0B', '#F97316', '#10B981', '#64748B']
    total = sum(counts)

    fig, ax = plt.subplots(figsize=(13.2, 7.8), dpi=140)
    fig.patch.set_facecolor('#F8FAFC')
    ax.set_facecolor('#F8FAFC')

    if total == 0:
        ax.text(0.5, 0.5, 'No ticket data available', ha='center', va='center',
                fontsize=13, color='#64748B', fontweight='semibold')
        ax.axis('off')
    else:
        percentages = [(count / total * 100) if total else 0 for count in counts]
        positions = range(len(statuses))
        bars = ax.barh(
            positions,
            percentages,
            color=colors,
            edgecolor='white',
            linewidth=1.2,
            height=0.62,
        )
        ax.set_yticks(list(positions), labels=statuses)
        ax.invert_yaxis()
        ax.set_xlim(0, max(100, max(percentages) + 12))
        ax.set_xlabel('Share of Tickets (%)', color='#334155', fontsize=12, fontweight='bold', labelpad=10)
        ax.tick_params(axis='x', labelsize=11, colors='#475569')
        ax.tick_params(axis='y', labelsize=12, colors='#0F172A')
        ax.grid(axis='x', color='#E2E8F0', linewidth=1.0, alpha=0.95)
        ax.set_axisbelow(True)
        for spine in ['top', 'right']:
            ax.spines[spine].set_visible(False)
        ax.spines['left'].set_color('#CBD5E1')
        ax.spines['bottom'].set_color('#CBD5E1')

        for bar, count, percentage in zip(bars, counts, percentages):
            ax.text(
                min(bar.get_width() + 1.2, ax.get_xlim()[1] - 1),
                bar.get_y() + bar.get_height() / 2,
                f'{count} ({percentage:.1f}%)',
                va='center',
                ha='left',
                fontsize=11,
                color='#0F172A',
                fontweight='bold'
            )
    ax.set_title('Ticket Status Distribution', fontsize=17, color='#0F172A', pad=12, weight='bold')

    buffer = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buffer, format='png', bbox_inches='tight', transparent=False)
    plt.close(fig)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='image/png')


@admin_required
def pie_chart(request):
    statuses = ['Open', 'In Progress', 'Reopen', 'Resolved', 'Closed']
    counts = [TicketDetail.objects.filter(TICKET_STATUS=s).count() for s in statuses]
    colors = ['#3B82F6', '#F59E0B', '#F97316', '#10B981', '#64748B']
    total = sum(counts)
    percentages = [round((count / total * 100), 1) if total else 0 for count in counts]
    return render(request, 'ticket_status_pie.html', {
        'status_labels': json.dumps(statuses),
        'status_counts': json.dumps(counts),
        'status_colors': json.dumps(colors),
        'status_percentages': json.dumps(percentages),
        'status_total': total,
    })


@admin_required
def Bar_chart(request):
    statuses = ['Open', 'In Progress', 'Reopen', 'Resolved', 'Closed']
    counts = [TicketDetail.objects.filter(TICKET_STATUS=s).count() for s in statuses]
    colors = ['#3B82F6', '#F59E0B', '#F97316', '#10B981', '#64748B']
    status_labels = ['Open', 'In\nProgress', 'Reopen', 'Resolved', 'Closed']

    fig, ax = plt.subplots(figsize=(13.4, 8.0), dpi=140)
    fig.patch.set_facecolor('#F8FAFC')
    ax.set_facecolor('#F8FAFC')

    bars = ax.bar(status_labels, counts, color=colors, edgecolor='white', linewidth=1.0, width=0.62)
    ax.set_title('Ticket Status Count', fontsize=17, color='#0F172A', pad=12, weight='bold')
    ax.set_ylabel('Tickets', color='#334155', fontsize=12, fontweight='bold')
    ax.tick_params(axis='x', rotation=0, labelsize=11.5, colors='#334155', pad=10)
    ax.tick_params(axis='y', labelsize=11, colors='#475569')
    ax.grid(axis='y', color='#E2E8F0', linewidth=1.0, alpha=0.95)
    ax.set_axisbelow(True)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    ax.spines['left'].set_color('#CBD5E1')
    ax.spines['bottom'].set_color('#CBD5E1')
    upper_limit = max(counts + [1])
    ax.set_ylim(0, upper_limit + max(1, upper_limit * 0.22))

    for bar, value in zip(bars, counts):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(0.12, upper_limit * 0.04),
            str(value),
            ha='center',
            va='bottom',
            fontsize=11,
            color='#0F172A',
            fontweight='bold'
        )

    if sum(counts) == 0:
        ax.text(0.5, 0.5, 'No ticket data available', transform=ax.transAxes,
                ha='center', va='center', fontsize=12, color='#64748B', fontweight='semibold')

    buffer = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buffer, format='png', bbox_inches='tight', transparent=False)
    plt.close(fig)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='image/png')

@login_required
def comment_view(request, pk, action):
    ticket   = get_object_or_404(TicketDetail, id=pk)
    action = action.lower()
    if action not in ['closing_comment', 'reopen_comment']:
        messages.error(request, "Only close note and reopen reason are allowed.")
        return redirect('ticketinfo', pk=pk)
    is_admin = _is_admin_user(request.user)
    if is_admin and action == 'closing_comment':
        messages.error(request, "Admins cannot perform close flow. You can delete the ticket if needed.")
        return redirect('ticketinfo', pk=pk)
    if action == 'reopen_comment' and not (is_admin or ticket.TICKET_CREATED_id == request.user.id):
        messages.error(request, "Only the ticket creator or admin can add reopen comment.")
        return redirect('ticketinfo', pk=pk)
    if (
        action == 'closing_comment'
        and _is_close_locked_by_rejection(request.user, ticket)
    ):
        messages.error(request, "You have already rejected this ticket and cannot close it now.")
        return redirect('ticketinfo', pk=pk)
    if (
        action == 'reopen_comment'
        and not is_admin
        and TicketHistory.objects.filter(
            ticket=ticket,
            action_type='REOPENED',
            changed_by=request.user,
        ).exists()
    ):
        messages.error(request, "You can reopen a ticket only once. Please create a new ticket.")
        return redirect('ticketinfo', pk=pk)
    if action == 'closing_comment' and not _can_user_close_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to close this ticket.")
        return redirect('ticketinfo', pk=pk)
    if action == 'reopen_comment' and not is_admin and not _can_work_on_ticket(request.user, ticket):
        messages.error(request, "You do not have permission to comment on this ticket.")
        return redirect('ticketinfo', pk=pk)

    is_dept_member = is_senior_dept_member = False
    if ticket.assigned_department:
        membership = DepartmentMember.objects.filter(
            user=request.user, department=ticket.assigned_department, is_active=True
        ).first()
        if membership:
            is_dept_member = True
            is_senior_dept_member = membership.role in ('SENIOR_MEMBER', 'LEAD', 'MANAGER')

    is_agent = is_admin or is_senior_dept_member

    if request.method == 'POST':
        form = UserCommentForm(request.POST, request.FILES)
        if form.is_valid():
            comment          = form.save(commit=False)
            comment.ticket     = ticket
            comment.user     = request.user
            comment.save()

            canned_id = request.POST.get('canned_response_id')
            if canned_id:
                try:
                    CannedResponse.objects.get(id=canned_id).increment_usage()
                except CannedResponse.DoesNotExist:
                    pass

            TicketHistory.objects.create(
                ticket=ticket, changed_by=request.user,
                action_type='COMMENTED',
                description=f'Comment added by {request.user.username}',
            )
            log_activity(
                request.user, 'COMMENTED',
                f'Commented on ticket: {ticket.TICKET_TITLE}', ticket=ticket,
            )

            notify_ticket_commented(
                ticket, request.user,
                comment.Closing_comment or comment.Reopen_comment or ''
            )

            if action == 'closing_comment':
                return CloseTicket(request, pk)
            elif action == 'reopen_comment':
                return reopenticket(request, pk)
    else:
        form = UserCommentForm()
        if action == 'closing_comment':
            form.fields['Reopen_comment'].widget = forms.HiddenInput()
        elif action == 'reopen_comment':
            form.fields['Closing_comment'].widget = forms.HiddenInput()

    canned_responses = []
    if is_agent:
        canned_responses = CannedResponse.objects.filter(is_active=True).filter(
            Q(is_public=True) | Q(department=ticket.assigned_department)
        )

    return render(request, 'Comment.html', {
        'form':             form,
        'ticket':             ticket,
        'action':           action,
        'is_admin':         is_admin,
        'is_agent':         is_agent,
        'canned_responses': canned_responses,
    })


@login_required
def download_file(request, pk):
    text_file = get_object_or_404(UserComment, id=pk)
    if not _can_view_ticket(request.user, text_file.ticket):
        messages.error(request, "You do not have permission to access this attachment.")
        return redirect('base')
    response  = FileResponse(open(text_file.TextFile.path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="{text_file.TextFile.name}"'
    return response

@admin_required
def category_list(request):
    categories = Category.objects.all()
    for cat in categories:
        cat.keywords_list = [k.strip() for k in cat.ml_keywords.split(',')] if cat.ml_keywords else []
    return render(request, 'admin/category_list.html', {'categories': categories})


@admin_required
def category_create(request):
    form = CategoryForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Category created.")
        return redirect('category_list')
    return render(request, 'admin/category_form.html', {'form': form})


@admin_required
def category_edit(request, pk):
    category = get_object_or_404(Category, id=pk)
    form     = CategoryForm(request.POST or None, instance=category)
    if form.is_valid():
        form.save()
        messages.success(request, f'Category "{category.name}" updated.')
        return redirect('category_list')
    return render(request, 'admin/category_form.html', {'form': form})


@admin_required
def category_delete(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('category_list')
    category = get_object_or_404(Category, id=pk)
    name     = category.name
    category.delete()
    messages.success(request, f'Category "{name}" deleted.')
    return redirect('category_list')

@login_required
def advanced_dashboard(request):
    visible_tickets = _exclude_inactive_department_tickets(TicketDetail.objects.all())
    context = {
        'total_tickets':    visible_tickets.count(),
        'open_tickets':     visible_tickets.filter(TICKET_STATUS='Open').count(),
        'in_progress':    visible_tickets.filter(TICKET_STATUS='In Progress').count(),
        'closed_tickets':   visible_tickets.filter(TICKET_STATUS='Closed').count(),
        'my_tickets':       visible_tickets.filter(TICKET_CREATED=request.user).count(),
        'urgent_tickets':   visible_tickets.filter(priority='URGENT').count(),
        'high_priority':  visible_tickets.filter(priority='HIGH').count(),
        'medium_priority':visible_tickets.filter(priority='MEDIUM').count(),
        'low_priority':   visible_tickets.filter(priority='LOW').count(),
        'recent_tickets':   visible_tickets.order_by('-TICKET_CREATED_ON')[:5],
    }
    return render(request, 'dashboard/advanced.html', context)

@login_required
@department_member_required
def department_dashboard(request, dept_id=None):
    if dept_id:
        department = get_object_or_404(Department, id=dept_id)
        if not request.user.is_superuser:
            if not department.is_active:
                messages.error(request, 'This department is inactive.')
                return redirect('base')
            if not DepartmentMember.objects.filter(
                user=request.user, department=department, is_active=True
            ).exists():
                messages.error(request, 'You are not a member of this department.')
                return redirect('base')
    else:
        membership = DepartmentMember.objects.filter(
            user=request.user,
            is_active=True,
            department__is_active=True,
        ).first()
        if not membership:
            messages.error(request, 'You are not a member of any department.')
            return redirect('base')
        department = membership.department

    tickets   = TicketDetail.objects.filter(assigned_department=department)\
                .select_related('TICKET_CREATED', 'category', 'assigned_to')
    members = DepartmentMember.objects.filter(department=department, is_active=True)\
                .select_related('user')

    stats = {
        'total':      tickets.count(),
        'open':       tickets.filter(TICKET_STATUS='Open').count(),
        'in_progress':tickets.filter(TICKET_STATUS='In Progress').count(),
        'closed':     tickets.filter(TICKET_STATUS='Closed').count(),
        'resolved':   tickets.filter(TICKET_STATUS='Resolved').count(),
        'members':    members.count(),
    }
    user_membership = members.filter(user=request.user).first()
    return render(request, 'department_dashboard.html', {
        'department':      department,
        'tickets':           tickets[:20],
        'members':         members,
        'stats':           stats,
        'user_role':       user_membership.role if user_membership else None,
        'user_permissions':user_membership if user_membership else None,
    })


@login_required
def department_members(request, dept_id=None):
    user_departments = Department.objects.filter(
        departmentmember__user=request.user,
        departmentmember__is_active=True,
        is_active=True,
    ).distinct().order_by('name')

    if request.user.is_superuser:
        if dept_id is not None:
            scoped_departments = Department.objects.filter(id=dept_id, is_active=True)
        else:
            scoped_departments = Department.objects.filter(is_active=True).order_by('name')
    else:
        if not user_departments.exists():
            messages.error(request, 'You are not a member of any department.')
            return redirect('base')
        if dept_id is not None:
            scoped_departments = user_departments.filter(id=dept_id)
            if not scoped_departments.exists():
                messages.error(request, 'You are not a member of this department.')
                return redirect('department_members')
        else:
            scoped_departments = user_departments

    department = scoped_departments.first()
    scoped_department_ids = list(scoped_departments.values_list('id', flat=True))

    members = DepartmentMember.objects.filter(
        department_id__in=scoped_department_ids, is_active=True
    ).select_related('user', 'user__userprofile').order_by('department__name', 'role', 'user__username')

    dept_tickets = TicketDetail.objects.filter(
        assigned_department_id__in=scoped_department_ids
    ).select_related('assigned_to', 'category')
    open_statuses = ['Open', 'In Progress', 'Reopen']
    today = timezone.localdate()

    dept_stats = {
        'total': dept_tickets.count(),
        'open': dept_tickets.filter(TICKET_STATUS='Open').count(),
        'in_progress': dept_tickets.filter(TICKET_STATUS='In Progress').count(),
        'resolved': dept_tickets.filter(TICKET_STATUS__in=['Closed', 'Resolved']).count(),
        'overdue': dept_tickets.filter(TICKET_DUE_DATE__lt=today).exclude(
            TICKET_STATUS__in=['Closed', 'Resolved']
        ).count(),
    }

    def _build_member_stats_for_scope(scope_members, scope_tickets):
        resolved_map = {
            row['assigned_to']: row['total']
            for row in scope_tickets.filter(TICKET_STATUS__in=['Closed', 'Resolved'])
            .values('assigned_to').annotate(total=Count('id'))
        }
        active_map = {
            row['assigned_to']: row['total']
            for row in scope_tickets.filter(TICKET_STATUS__in=open_statuses)
            .values('assigned_to').annotate(total=Count('id'))
        }
        overdue_map = {
            row['assigned_to']: row['total']
            for row in scope_tickets.filter(TICKET_DUE_DATE__lt=today)
            .exclude(TICKET_STATUS__in=['Closed', 'Resolved'])
            .values('assigned_to').annotate(total=Count('id'))
        }

        stats = []
        for m in scope_members:
            resolved = resolved_map.get(m.user_id, 0)
            active = active_map.get(m.user_id, 0)
            overdue = overdue_map.get(m.user_id, 0)
            total_worked = resolved + active
            completion_rate = int((resolved / total_worked) * 100) if total_worked > 0 else 0

            stats.append({
                'membership':      m,
                'resolved':        resolved,
                'active':          active,
                'overdue':         overdue,
                'completion_rate': completion_rate,
                'profile_image':   '',
                'is_current_user': m.user_id == request.user.id,
            })
        stats.sort(key=lambda x: (-x['resolved'], x['active'], x['membership'].user.username))
        return stats

    member_stats = _build_member_stats_for_scope(members, dept_tickets)
    role_counts = {r['role']: r['total'] for r in members.values('role').annotate(total=Count('id'))}
    recent_tickets = dept_tickets.order_by('-updated_at')[:8]

    department_sections = []
    for d in scoped_departments:
        dept_members = members.filter(department=d)
        dept_member_count = dept_members.count()
        d_tickets = dept_tickets.filter(assigned_department=d)
        d_role_counts = {r['role']: r['total'] for r in dept_members.values('role').annotate(total=Count('id'))}
        department_sections.append({
            'department': d,
            'total_members': dept_member_count,
            'member_stats': _build_member_stats_for_scope(dept_members, d_tickets),
            'role_counts': d_role_counts,
            'recent_tickets': d_tickets.order_by('-updated_at')[:8],
        })

    multi_departments = []
    for user_dept in user_departments:
        user_dept_tickets = TicketDetail.objects.filter(assigned_department=user_dept)
        multi_departments.append({
            'department': user_dept,
            'members': DepartmentMember.objects.filter(
                department=user_dept, is_active=True
            ).count(),
            'total': user_dept_tickets.count(),
            'open': user_dept_tickets.filter(
                TICKET_STATUS__in=['Open', 'In Progress', 'Reopen']
            ).count(),
            'resolved': user_dept_tickets.filter(
                TICKET_STATUS__in=['Closed', 'Resolved']
            ).count(),
        })

    return render(request, 'department_members.html', {
        'department':      department,
        'member_stats':    member_stats,
        'total_members':   members.count(),
        'dept_stats':      dept_stats,
        'recent_tickets':  recent_tickets,
        'role_counts':     role_counts,
        'multi_departments': multi_departments,
        'department_sections': department_sections,
        'is_multi_department_view': dept_id is None and len(scoped_department_ids) > 1,
        'department_display_name': ", ".join(
            scoped_departments.values_list('name', flat=True)
        ) if dept_id is None and len(scoped_department_ids) > 1 else (department.name if department else "Departments"),
    })

@admin_required
def admin_department_list(request):
    return _render_admin_department_list(request)


def _render_admin_department_list(request, create_form=None, edit_forms=None):
    departments = Department.objects.all().order_by('-is_active', 'name')
    inactive_departments = departments.filter(is_active=False)
    all_users = User.objects.filter(is_active=True, is_superuser=False).order_by('username')
    create_form = create_form or DepartmentAdminForm(prefix='create')
    edit_forms = edit_forms or {}
    dept_data = []
    total_members = 0
    total_open_tickets = 0

    for dept in departments:
        members = DepartmentMember.objects.filter(
            department=dept, is_active=True, user__is_superuser=False
        ).select_related('user')
        tickets = TicketDetail.objects.filter(assigned_department=dept)
        creator_user_ids = DepartmentMember.objects.filter(department=dept).values('user_id')
        created_tickets = TicketDetail.objects.filter(
            TICKET_CREATED_id__in=creator_user_ids
        ).select_related('TICKET_CREATED', 'assigned_department').order_by('-TICKET_CREATED_ON', '-id')
        assigned_tickets = tickets.select_related('TICKET_CREATED', 'assigned_department').order_by('-TICKET_CREATED_ON', '-id')
        open_tickets = tickets.exclude(TICKET_STATUS__in=['Closed', 'Resolved'])
        member_count = members.count()
        open_count = open_tickets.count()
        if dept.is_active:
            total_members += member_count
            total_open_tickets += open_count
        dept_data.append({
            'department': dept,
            'members':    members,
            'total':      tickets.count(),
            'open':       tickets.filter(TICKET_STATUS='Open').count(),
            'resolved':   tickets.filter(TICKET_STATUS__in=['Closed', 'Resolved']).count(),
            'member_count': member_count,
            'active_tickets': open_count,
            'created_ticket_count': created_tickets.count(),
            'assigned_ticket_count': assigned_tickets.count(),
            'created_tickets': created_tickets[:5],
            'assigned_tickets': assigned_tickets[:5],
            'edit_form': edit_forms.get(dept.id, DepartmentAdminForm(instance=dept, prefix=f'dept-{dept.id}')),
            'member_form': DepartmentMemberForm(prefix=f'member-{dept.id}') if dept.is_active else None,
        })

    return render(request, 'admin_department_list.html', {
        'dept_data':    dept_data,
        'all_users':    all_users,
        'departments':  departments,
        'create_form':  create_form,
        'department_count': departments.count(),
        'member_count': total_members,
        'open_ticket_count': total_open_tickets,
        'inactive_departments': inactive_departments,
    })


@admin_required
def admin_create_department(request):
    if request.method != 'POST':
        return redirect('admin_department_list')

    raw_name = (request.POST.get('create-name') or '').strip()
    raw_code = (request.POST.get('create-code') or '').strip().upper()
    inactive_name_match = Department.objects.filter(is_active=False, name__iexact=raw_name).first() if raw_name else None
    inactive_code_match = Department.objects.filter(is_active=False, code__iexact=raw_code).first() if raw_code else None

    if inactive_name_match and inactive_code_match and inactive_name_match.id != inactive_code_match.id:
        form = DepartmentAdminForm(request.POST, prefix='create')
        form.add_error(None, 'That department name and code belong to different deleted departments. Use the original pair or edit an existing department.')
        messages.error(request, 'Please correct the new department details below.')
        return _render_admin_department_list(request, create_form=form)

    reactivated_department = inactive_name_match or inactive_code_match
    form = DepartmentAdminForm(request.POST, instance=reactivated_department, prefix='create')
    if not form.is_valid():
        messages.error(request, 'Please correct the new department details below.')
        return _render_admin_department_list(request, create_form=form)

    department = form.save(commit=False)
    if not department.created_by_id:
        department.created_by = request.user
    if reactivated_department:
        department.is_active = True
    department.save()
    if reactivated_department:
        _restore_department_ticket_assignments(department, changed_by=request.user)
        log_activity(request.user, 'UPDATED', f'Reactivated department: {department.name}')
        messages.success(request, f'{department.name} department reactivated successfully.')
    else:
        log_activity(request.user, 'CREATED', f'Created department: {department.name}')
        messages.success(request, f'{department.name} department created successfully.')
    return redirect('admin_department_list')


@admin_required
def admin_update_department(request, dept_id):
    if request.method != 'POST':
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id, is_active=True)
    form = DepartmentAdminForm(request.POST, instance=department, prefix=f'dept-{dept_id}')
    if not form.is_valid():
        messages.error(request, f'Please correct the details for {department.name}.')
        return _render_admin_department_list(request, edit_forms={dept_id: form})

    updated_department = form.save()
    log_activity(request.user, 'UPDATED', f'Updated department: {updated_department.name}')
    messages.success(request, f'{updated_department.name} updated successfully.')
    return redirect('admin_department_list')


@admin_required
def admin_delete_department(request, dept_id):
    if request.method != 'POST':
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id, is_active=True)
    affected_tickets = TicketDetail.objects.filter(assigned_department=department)
    affected_ticket_ids = list(affected_tickets.values_list('id', flat=True))
    assigned_ticket_rows = list(
        affected_tickets.filter(assigned_to__isnull=False).values_list('id', 'assigned_to_id', 'assigned_to__username')
    )
    for ticket_id, assigned_to_id, assigned_to_username in assigned_ticket_rows:
        TicketHistory.objects.create(
            ticket_id=ticket_id,
            changed_by=request.user,
            action_type='UPDATED',
            field_name='department_inactivation_assignee',
            old_value=str(assigned_to_id or ''),
            new_value='',
            description=(
                f'Stored {assigned_to_username or "previous assignee"} before '
                f'{department.name} department was marked inactive.'
            ),
        )
    affected_tickets.update(
        assigned_to=None,
        TICKET_HOLDER='',
        assignment_type='UNASSIGNED',
        assigned_by=None,
        assigned_at=None,
    )
    if affected_ticket_ids:
        MyCart.objects.filter(ticket_id__in=affected_ticket_ids).delete()
    department.is_active = False
    department.save(update_fields=['is_active'])

    log_activity(request.user, 'UPDATED', f'Inactivated department: {department.name}')
    messages.success(request, f'{department.name} department marked inactive successfully.')
    return redirect('admin_department_list')


@admin_required
def admin_reactivate_department(request, dept_id):
    if request.method != 'POST':
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id, is_active=False)
    department.is_active = True
    department.save(update_fields=['is_active'])
    _restore_department_ticket_assignments(department, changed_by=request.user)
    log_activity(request.user, 'UPDATED', f'Reactivated department: {department.name}')
    messages.success(request, f'{department.name} department reactivated successfully.')
    return redirect('admin_department_list')


@admin_required
def admin_permanently_delete_inactive_department(request):
    if request.method != 'POST':
        return redirect('admin_department_list')

    dept_id = request.POST.get('inactive_department_id')
    if not dept_id:
        messages.error(request, 'Please select an inactive department to delete.')
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id, is_active=False)
    DepartmentMember.objects.filter(department=department).delete()
    department_name = department.name
    department.delete()
    log_activity(request.user, 'DELETED', f'Permanently deleted department: {department_name}')
    messages.success(request, f'{department_name} department permanently deleted successfully.')
    return redirect('admin_department_list')


@admin_required
def admin_department_tickets(request, dept_id):
    department = get_object_or_404(Department, id=dept_id)
    ticket_view = (request.GET.get('view') or 'assigned').strip().lower()
    if ticket_view not in {'assigned', 'created'}:
        ticket_view = 'assigned'

    member_user_ids = DepartmentMember.objects.filter(
        department=department
    ).values_list('user_id', flat=True)

    assigned_qs = TicketDetail.objects.filter(
        assigned_department=department
    ).select_related('TICKET_CREATED', 'assigned_department', 'assigned_to', 'category')
    created_qs = TicketDetail.objects.filter(
        TICKET_CREATED_id__in=member_user_ids
    ).select_related('TICKET_CREATED', 'assigned_department', 'assigned_to', 'category')

    q = (request.GET.get('q') or '').strip()
    if q:
        assigned_filters = (
            Q(TICKET_TITLE__icontains=q) |
            Q(TICKET_DESCRIPTION__icontains=q) |
            Q(TICKET_CREATED__username__icontains=q)
        )
        created_filters = (
            Q(TICKET_TITLE__icontains=q) |
            Q(TICKET_DESCRIPTION__icontains=q) |
            Q(TICKET_CREATED__username__icontains=q)
        )
        if q.isdigit():
            assigned_filters |= Q(id=int(q))
            created_filters |= Q(id=int(q))
        assigned_qs = assigned_qs.filter(assigned_filters)
        created_qs = created_qs.filter(created_filters)

    status = (request.GET.get('status') or '').strip()
    if status:
        assigned_qs = assigned_qs.filter(TICKET_STATUS=status)
        created_qs = created_qs.filter(TICKET_STATUS=status)

    selected_qs = assigned_qs if ticket_view == 'assigned' else created_qs
    paginator = Paginator(selected_qs.order_by('-TICKET_CREATED_ON', '-id'), 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    query = request.GET.copy()
    query.pop('page', None)

    return render(request, 'admin_department_tickets.html', {
        'department': department,
        'ticket_view': ticket_view,
        'tickets': page_obj,
        'assigned_count': assigned_qs.count(),
        'created_count': created_qs.count(),
        'search_query': q,
        'status_filter': status,
        'pagination_query': query.urlencode(),
        'status_choices': TicketDetail.choice,
    })


def _apply_department_role(membership, role):
    role_permissions = ROLE_PERMISSION_MATRIX.get(role, ROLE_PERMISSION_MATRIX['MEMBER'])
    membership.role = role
    membership.is_active = True
    membership.can_close_tickets = role_permissions['can_close_tickets']
    membership.can_assign_tickets = role_permissions['can_assign_tickets']
    membership.can_delete_tickets = role_permissions['can_delete_tickets']
    return membership


@admin_required
def admin_add_member(request, dept_id):
    department = get_object_or_404(Department, id=dept_id, is_active=True)

    if request.method == 'POST':
        form = DepartmentMemberForm(request.POST, prefix=f'member-{dept_id}')
        if not form.is_valid():
            messages.error(request, 'Please select a valid user and role.')
            return redirect('admin_department_list')

        user = form.cleaned_data['user_id']
        role = form.cleaned_data['role']
        membership, created = DepartmentMember.objects.get_or_create(
            user=user,
            department=department,
            defaults={'added_by': request.user},
        )
        previous_role = membership.role
        was_active = membership.is_active if not created else False
        _apply_department_role(membership, role)
        if created:
            membership.added_by = request.user
            membership.save()
            status_message = f'{user.username} added to {department.name}.'
            level = 'success'
        else:
            membership.save(update_fields=[
                'role', 'is_active', 'can_close_tickets',
                'can_assign_tickets', 'can_delete_tickets',
            ])
            if was_active and previous_role == role:
                status_message = f'{user.username} is already in {department.name}.'
                level = 'info'
            elif was_active:
                status_message = f'{user.username} role updated in {department.name}.'
                level = 'success'
            else:
                status_message = f'{user.username} added to {department.name} successfully.'
                level = 'success'

        rejected_ticket_ids = TicketHistory.objects.filter(
            changed_by=user,
            action_type='REJECTED',
        ).values_list('ticket_id', flat=True)

        for ticket in TicketDetail.objects.filter(
            assigned_department=department,
        ).exclude(TICKET_STATUS__in=['Closed', 'Resolved']).exclude(id__in=rejected_ticket_ids):
            MyCart.objects.get_or_create(user=user, ticket=ticket)

        log_activity(
            request.user,
            'UPDATED',
            f'Added {user.username} to {department.name} as {role}'
        )
        getattr(messages, level)(request, status_message)

    return redirect('admin_department_list')


@admin_required
def admin_update_member_role(request, dept_id, user_id):
    if request.method != 'POST':
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id, is_active=True)
    user = get_object_or_404(User, id=user_id, is_superuser=False)
    membership = get_object_or_404(DepartmentMember, department=department, user=user)
    role = request.POST.get('role', membership.role)
    if role not in ROLE_PERMISSION_MATRIX:
        messages.error(request, 'Invalid department role selected.')
        return redirect('admin_department_list')

    _apply_department_role(membership, role)
    membership.save(update_fields=[
        'role', 'is_active', 'can_close_tickets',
        'can_assign_tickets', 'can_delete_tickets',
    ])
    log_activity(request.user, 'UPDATED', f'Updated {user.username} role in {department.name} to {role}')
    messages.success(request, f'{user.username} is now {membership.display_role} in {department.name}.')
    return redirect('admin_department_list')


@admin_required
def admin_remove_member(request, dept_id, user_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('admin_department_list')

    department = get_object_or_404(Department, id=dept_id)
    user       = get_object_or_404(User, id=user_id)

    DepartmentMember.objects.filter(user=user, department=department).update(is_active=False)

    dept_ticket_ids = TicketDetail.objects.filter(
        assigned_department=department
    ).values_list('id', flat=True)
    MyCart.objects.filter(user=user, ticket_id__in=dept_ticket_ids).delete()

    log_activity(request.user, 'UPDATED',
                 f'Removed {user.username} from {department.name}')
    messages.success(request, f'{user.username} removed from {department.name}.')
    return redirect('admin_department_list')

@login_required
def notifications_list(request):
    notifs      = _notifications_for_user(request.user)
    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'unread':
        notifs = notifs.filter(is_read=False)
    elif filter_type == 'read':
        notifs = notifs.filter(is_read=True)
    paginator = Paginator(notifs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    all_notifs = _notifications_for_user(request.user)
    stats = {
        'total':  all_notifs.count(),
        'unread': all_notifs.filter(is_read=False).count(),
        'read':   all_notifs.filter(is_read=True).count(),
    }
    return render(request, 'notifications_list.html', {
        'notifications': page_obj, 'stats': stats, 'filter_type': filter_type,
    })


@login_required
def mark_notification_read(request, notification_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('notifications_list')
    notif    = get_object_or_404(_notifications_for_user(request.user), id=notification_id)
    notif.mark_as_read()
    return _redirect_to_safe_next(request, 'notifications_list')


@login_required
def mark_all_read(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('notifications_list')
    updated = _notifications_for_user(request.user).filter(is_read=False).update(is_read=True)
    messages.success(request, f'Marked {updated} notifications as read.')
    return _redirect_to_safe_next(request, 'notifications_list')


@login_required
def delete_notification(request, notification_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('notifications_list')
    notif = get_object_or_404(_notifications_for_user(request.user), id=notification_id)
    notif.delete()
    messages.success(request, 'Notification deleted.')
    return _redirect_to_safe_next(request, 'notifications_list')


@login_required
def delete_all_notifications(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('notifications_list')
    deleted, _ = _notifications_for_user(request.user).delete()
    messages.success(request, f'Deleted {deleted} notifications.')
    return _redirect_to_safe_next(request, 'notifications_list')


@login_required
def notification_count_api(request):
    count = _notifications_for_user(request.user).filter(is_read=False).count()
    return JsonResponse({'count': count})

@admin_required
def analytics_dashboard(request):
    def _department_from_query(param_name):
        dept_id = request.GET.get(param_name)
        if not dept_id:
            return None
        try:
            return Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return None

    range_type = request.GET.get('range', '30_days')
    start_date, end_date = get_date_range(range_type)

    if range_type == 'custom':
        try:
            start_date = datetime.strptime(request.GET.get('start_date', ''), '%Y-%m-%d').date()
            end_date   = datetime.strptime(request.GET.get('end_date',   ''), '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format.')
            start_date, end_date = get_date_range('30_days')

    department = _department_from_query('department')
    resolver_department = _department_from_query('resolver_department') or department
    creator_department = _department_from_query('creator_department') or department

    context = {
        'start_date':           start_date,
        'end_date':             end_date,
        'range_type':           range_type,
        'selected_department':  department,
        'selected_resolver_department': resolver_department,
        'selected_creator_department': creator_department,
        'departments':          Department.objects.filter(is_active=True),
        'stats':                get_ticket_statistics(start_date, end_date, department),
        'dept_stats':           get_department_statistics(start_date, end_date),
        'dept_comparison':      get_department_comparison(),
        'top_creators':         get_top_ticket_creators(5, start_date, end_date, creator_department),
        'top_resolvers':        get_top_ticket_resolvers(5, start_date, end_date, resolver_department),
        'priority_dist':        get_priority_distribution(start_date, end_date, department),
        'category_dist':        get_category_distribution(start_date, end_date),
    }
    return render(request, 'analytics_dashboard.html', context)


@admin_required
def api_tickets_over_time(request):
    try:
        range_type       = request.GET.get('range', '30_days')
        start_date, end_date = get_date_range(range_type)
        dept_id          = request.GET.get('department')
        department       = Department.objects.get(id=dept_id) if dept_id else None
        data             = get_tickets_over_time(start_date, end_date, department)
        return JsonResponse({'labels': [i['date'] for i in data], 'data': [i['count'] for i in data]})
    except Exception as e:
        logger.exception("Failed to build tickets-over-time API response.")
        return JsonResponse({'error': 'Unable to load chart data right now.', 'labels': [], 'data': []}, status=500)


@admin_required
def api_department_comparison(request):
    try:
        return JsonResponse(get_department_comparison())
    except Exception as e:
        logger.exception("Failed to build department comparison API response.")
        return JsonResponse({'error': 'Unable to load comparison data right now.'}, status=500)


@admin_required
def api_priority_distribution(request):
    try:
        range_type       = request.GET.get('range', '30_days')
        start_date, end_date = get_date_range(range_type)
        data             = get_priority_distribution(start_date, end_date)
        return JsonResponse({
            'labels': ['Urgent','High','Medium','Low'],
            'data':   [data['URGENT'],data['HIGH'],data['MEDIUM'],data['LOW']],
            'colors': ['#ef4444','#f59e0b','#3b82f6','#94a3b8'],
        })
    except Exception as e:
        logger.exception("Failed to build priority distribution API response.")
        return JsonResponse({'error': 'Unable to load priority data right now.'}, status=500)


@admin_required
def api_category_distribution(request):
    try:
        range_type       = request.GET.get('range', '30_days')
        start_date, end_date = get_date_range(range_type)
        data             = get_category_distribution(start_date, end_date)
        if not data:
            return JsonResponse({'labels':['No Categories'],'data':[1],'colors':['#94a3b8']})
        return JsonResponse({
            'labels': [i['name']  for i in data],
            'data':   [i['count'] for i in data],
            'colors': [i['color'] for i in data],
        })
    except Exception as e:
        logger.exception("Failed to build category distribution API response.")
        return JsonResponse({'error': 'Unable to load category data right now.'}, status=500)


@admin_required
def export_analytics_excel(request):
    import openpyxl
    from openpyxl.styles import Font

    def _department_from_query(param_name):
        dept_id = request.GET.get(param_name)
        if not dept_id:
            return None
        try:
            return Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return None

    range_type       = request.GET.get('range', '30_days')
    start_date, end_date = get_date_range(range_type)
    department = _department_from_query('department')
    resolver_department = _department_from_query('resolver_department') or department
    creator_department = _department_from_query('creator_department') or department
    data     = prepare_export_data(
        start_date,
        end_date,
        department=department,
        resolver_department=resolver_department,
        creator_department=creator_department,
    )

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1['A1'] = "Helpdesk Analytics Report"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A2'] = f"Period: {start_date} to {end_date}"
    ws1['A3'] = f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
    rows = [
        ('Total Tickets', data['statistics']['total']),
        ('Open',        data['statistics']['open']),
        ('In Progress', data['statistics']['in_progress']),
        ('Closed',      data['statistics']['closed']),
        ('Resolved',    data['statistics']['resolved']),
        ('Completion Rate', f"{data['statistics']['completion_rate']:.2f}%"),
        ('Avg Resolution (hrs)', data['statistics']['avg_resolution_hours']),
    ]
    for i, (label, val) in enumerate(rows, start=6):
        ws1.cell(row=i, column=1).value = label
        ws1.cell(row=i, column=2).value = val

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=analytics_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response

@login_required
def ticket_history(request, pk):
    if pk == 0:
        resolved_entry = TicketHistory.objects.filter(
            ticket=OuterRef('pk'),
            action_type__in=['RESOLVED', 'CLOSED']
        ).order_by('-changed_at')

        resolved_tickets = (
            TicketDetail.objects
            .filter(TICKET_STATUS__in=['Resolved', 'Closed'])
            .filter(
                Q(TICKET_CREATED=request.user) |
                Q(assigned_to=request.user) |
                Q(TICKET_CLOSED=request.user)
            )
            .annotate(
                resolved_by=Subquery(resolved_entry.values('changed_by__username')[:1]),
                resolved_on=Subquery(resolved_entry.values('changed_at')[:1]),
            )
            .select_related('assigned_department', 'assigned_to', 'TICKET_CREATED')
            .order_by('-resolved_on')
        )

        return render(request, 'ticket_history.html', {
            'resolved_history_mode': True,
            'resolved_tickets': resolved_tickets,
            'my_resolved_count': resolved_tickets.filter(TICKET_CLOSED=request.user).count(),
        })

    ticket = get_object_or_404(TicketDetail, id=pk)
    if not _can_view_ticket(request.user, ticket):
        messages.error(request, 'You do not have permission to view this ticket history.')
        return redirect('base')
    history = (
        TicketHistory.objects
        .filter(ticket=ticket)
        .select_related('changed_by')
        .order_by('-changed_at')
    )

    return render(request, 'ticket_history.html', {
        'ticket': ticket,
        'history': history
    })

@login_required
def rate_ticket(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if ticket.TICKET_CREATED != request.user:
        messages.error(request, 'Only the ticket creator can rate this ticket.')
        return redirect('ticketinfo', pk=pk)
    if ticket.TICKET_STATUS not in ['Closed', 'Resolved']:
        messages.error(request, 'You can only rate completed tickets.')
        return redirect('ticketinfo', pk=pk)
    if TicketRating.objects.filter(ticket=ticket).exists():
        messages.info(request, 'You have already rated this ticket.')
        return redirect('ticketinfo', pk=pk)

    if request.method == 'POST':
        form = TicketRatingForm(request.POST)
        if form.is_valid():
            rating          = form.save(commit=False)
            rating.ticket     = ticket
            rating.rated_by = request.user
            rating.save()
            notify_ticket_rated(ticket, rating)
            TicketHistory.objects.create(
                ticket=ticket, changed_by=request.user,
                action_type='UPDATED', field_name='rating',
                new_value=str(rating.rating),
                description=f'Ticket rated {rating.rating}⭐ by {request.user.username}',
            )
            messages.success(request, f'Thank you! You rated this ticket {rating.rating}⭐')
            return redirect('ticketinfo', pk=pk)
    else:
        form = TicketRatingForm()
    return render(request, 'rate_ticket.html', {'form': form, 'ticket': ticket})

@admin_required
def send_overdue_note(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if request.method != 'POST':
        return redirect('ticketinfo', pk=pk)

    if ticket.assigned_department_id and not getattr(ticket.assigned_department, 'is_active', False):
        messages.error(request, 'Overdue notes cannot be sent for inactive-department tickets.')
        return redirect('ticketinfo', pk=pk)

    if ticket.TICKET_STATUS in ['Closed', 'Resolved']:
        messages.error(request, 'This ticket is already completed.')
        return redirect('ticketinfo', pk=pk)

    if not ticket.is_overdue:
        messages.error(request, 'Overdue note can be sent only for overdue tickets.')
        return redirect('ticketinfo', pk=pk)

    note = request.POST.get('overdue_note', '').strip()
    if not note:
        messages.error(request, 'Please enter a note before sending.')
        return redirect('ticketinfo', pk=pk)

    recipient = User.objects.filter(id=ticket.assigned_to_id, is_active=True).first()

    sent_count = 0
    if recipient:
        create_notification(
            user=recipient,
            notification_type='TICKET_OVERDUE',
            title=f'Overdue reminder: #{ticket.id} {ticket.TICKET_TITLE}',
            message=note,
            ticket=ticket,
            extra_data={
                'sent_by': request.user.username,
                'is_admin_note': True,
            },
        )
        sent_count = 1

    if sent_count:
        TicketHistory.objects.create(
            ticket=ticket,
            changed_by=request.user,
            action_type='UPDATED',
            field_name='admin_overdue_note',
            description=f'Admin sent overdue reminder to {sent_count} users.',
            new_value=note,
        )
        log_activity(
            request.user,
            'UPDATED',
            f'Sent overdue reminder for ticket #{ticket.id}',
            description=note,
            ticket=ticket,
        )
        messages.success(request, f'Overdue reminder sent to {sent_count} users.')
    else:
        messages.warning(request, 'No eligible recipients found for this ticket.')

    return _ticketinfo_overdue_redirect(ticket)


@login_required
def reply_overdue_note(request, pk):
    ticket = get_object_or_404(TicketDetail, id=pk)
    if request.method != 'POST':
        return redirect('ticketinfo', pk=pk)

    if ticket.assigned_department_id and not getattr(ticket.assigned_department, 'is_active', False):
        messages.error(request, 'Overdue note replies are not allowed for inactive-department tickets.')
        return redirect('ticketinfo', pk=pk)

    if _is_admin_user(request.user):
        messages.error(request, 'Admin should use overdue note instead of member reply.')
        return redirect('ticketinfo', pk=pk)

    can_reply = ticket.assigned_to_id == request.user.id
    if not can_reply:
        messages.error(request, 'You are not allowed to reply to this overdue note.')
        return redirect('ticketinfo', pk=pk)

    reply_text = request.POST.get('overdue_note_reply', '').strip()
    if not reply_text:
        messages.error(request, 'Please enter a reply before sending.')
        return redirect('ticketinfo', pk=pk)

    TicketHistory.objects.create(
        ticket=ticket,
        changed_by=request.user,
        action_type='UPDATED',
        field_name='admin_overdue_note_reply',
        description=f'Overdue note reply by {request.user.username}.',
        new_value=reply_text,
    )
    log_activity(
        request.user,
        'COMMENTED',
        f'Replied on overdue note for ticket #{ticket.id}',
        description=reply_text,
        ticket=ticket,
    )

    for admin_user in User.objects.filter(is_superuser=True, is_active=True):
        create_notification(
            user=admin_user,
            notification_type='TICKET_COMMENTED',
            title=f'Overdue note reply: #{ticket.id} {ticket.TICKET_TITLE}',
            message=f'{request.user.username} wrote the overdue note for the ticket #{ticket.id}.',
            ticket=ticket,
            extra_data={
                'sent_by': request.user.username,
                'is_admin_note_reply': True,
            },
        )

    messages.success(request, 'Your reply was sent to admin.')
    return _ticketinfo_overdue_redirect(ticket)

@login_required
def department_analytics(request, dept_id):
    department = get_object_or_404(Department, id=dept_id)
    if not request.user.is_superuser:
        if not department.is_active:
            messages.error(request, 'This department is inactive.')
            return redirect('base')
        if not DepartmentMember.objects.filter(
            user=request.user,
            department=department,
            is_active=True,
            department__is_active=True,
        ).exists():
            messages.error(request, 'You are not allowed to view analytics for this department.')
            return redirect('base')

    range_type           = request.GET.get('range', '30_days')
    start_date, end_date = get_date_range(range_type)

    stats   = get_ticket_statistics(start_date, end_date, department)
    members = DepartmentMember.objects.filter(
        department=department, is_active=True
    ).select_related('user')

    member_user_ids = [m.user_id for m in members]

    created_map = {
        row['TICKET_CREATED']: row['total']
        for row in TicketDetail.objects.filter(
            TICKET_CREATED_id__in=member_user_ids,
            TICKET_CREATED_ON__range=(start_date, end_date)
        ).values('TICKET_CREATED').annotate(total=Count('id'))
    }
    resolved_map = {
        row['assigned_to']: row['total']
        for row in TicketDetail.objects.filter(
            assigned_to_id__in=member_user_ids,
            TICKET_STATUS__in=['Closed', 'Resolved'],
            TICKET_CLOSED_ON__range=(start_date, end_date)
        ).values('assigned_to').annotate(total=Count('id'))
    }
    active_map = {
        row['assigned_to']: row['total']
        for row in TicketDetail.objects.filter(
            assigned_to_id__in=member_user_ids,
            TICKET_STATUS__in=['Open', 'In Progress', 'Reopen']
        ).values('assigned_to').annotate(total=Count('id'))
    }

    member_stats = []
    for member in members:
        member_stats.append({
            'member':           member,
            'tickets_created':  created_map.get(member.user_id, 0),
            'tickets_resolved': resolved_map.get(member.user_id, 0),
            'active_tickets':   active_map.get(member.user_id, 0),
        })

    return render(request, 'department_dashboard.html', {
        'department':   department,
        'start_date':   start_date,
        'end_date':     end_date,
        'range_type':   range_type,
        'stats':        stats,
        'member_stats': member_stats,
    })

@admin_required
def export_analytics_pdf(request):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from io import BytesIO

    def _department_from_query(param_name):
        dept_id = request.GET.get(param_name)
        if not dept_id:
            return None
        try:
            return Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return None

    range_type           = request.GET.get('range', '30_days')
    start_date, end_date = get_date_range(range_type)
    department = _department_from_query('department')
    resolver_department = _department_from_query('resolver_department') or department
    creator_department = _department_from_query('creator_department') or department
    data     = prepare_export_data(
        start_date,
        end_date,
        department=department,
        resolver_department=resolver_department,
        creator_department=creator_department,
    )

    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
    )

    elements = [
        Paragraph("Helpdesk Analytics Report", title_style),
        Paragraph(
            f"Period: {start_date} to {end_date}<br/>"
            f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ),
        Spacer(1, 0.3 * inch),
    ]

    stats_data = [
        ['Metric', 'Value'],
        ['Total Tickets',          str(data['statistics']['total'])],
        ['Open',                 str(data['statistics']['open'])],
        ['In Progress',          str(data['statistics']['in_progress'])],
        ['Closed',               str(data['statistics']['closed'])],
        ['Resolved',             str(data['statistics']['resolved'])],
        ['Completion Rate',      f"{data['statistics']['completion_rate']:.2f}%"],
        ['Avg Resolution Time',  f"{data['statistics']['avg_resolution_hours']} hours"],
    ]

    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1,  0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',    (0, 0), (-1,  0), colors.whitesmoke),
        ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME',     (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1,  0), 13),
        ('BOTTOMPADDING',(0, 0), (-1,  0), 12),
        ('BACKGROUND',   (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID',         (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
    ]))

    elements.append(stats_table)

    if data.get('department_stats'):
        elements.append(Spacer(1, 0.4 * inch))
        elements.append(Paragraph("Department Breakdown", styles['Heading2']))
        elements.append(Spacer(1, 0.2 * inch))

        dept_rows = [['Department', 'Total', 'Open', 'Closed', 'Completion %', 'Members']]
        for dept in data['department_stats']:
            dept_rows.append([
                dept['name'],
                str(dept['total_tickets']),
                str(dept['open_tickets']),
                str(dept['closed_tickets']),
                f"{dept['completion_rate']:.1f}%",
                str(dept['members_count']),
            ])

        dept_table = Table(dept_rows)
        dept_table.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1,  0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR',    (0, 0), (-1,  0), colors.whitesmoke),
            ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME',     (0, 0), (-1,  0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1,  0), 11),
            ('BOTTOMPADDING',(0, 0), (-1,  0), 10),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        elements.append(dept_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename=analytics_{start_date}_{end_date}.pdf'
    )
    return response
